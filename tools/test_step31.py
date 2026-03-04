#!/usr/bin/env python
"""
Test script for Excel MVP Skeleton (Step 31)
Verifica que el archivo generado cumpla con los requisitos mínimos
"""

import os
from pathlib import Path
from openpyxl import load_workbook

# Configuration
OUTPUT_FILE = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

# Expected sheets (in order)
# Note: As of Step 33, 00_Home is expected to be first sheet
EXPECTED_SHEETS = [
    "00_Home",
    "01_Recetas",
    "02_RecetaVersiones",
    "03_Lotes",
    "04_LoteMediciones",
    "05_ItemsInventario",
    "06_MovimientosInventario",
    "07_Proveedores",
    "08_LotesInsumo",
    "09_ConsumosLote",
    "10_Productos",
    "11_Clientes",
    "12_Ventas",
    "13_VentasLineas",
    "14_Pagos",
    "15_FulfillmentVentaLote",
    "99_Listas",
    "Catalogos",
    "98_Ayuda"
]

# Expected headers for verification
EXPECTED_HEADERS = {
    "01_Recetas": ["receta_id", "nombre", "estado", "fecha_creacion", "notas"],
    "03_Lotes": ["batch_id", "recipe_version_id", "fecha_inicio", "fecha_coccion",
                "fecha_fermentacion", "fecha_maduracion", "fecha_envasado",
                "volumen_litros", "og_medido", "fg_medido", "abv_medido",
                "estado", "costo_real_batch", "notas"],
    "99_Listas": ["LIST_RECETAS_ESTADO"]
}

# Expected defined names
EXPECTED_DEFINED_NAMES = [
    "LIST_RECETAS_ESTADO",
    "LIST_LOTES_ESTADO",
    "LIST_ITEMS_TIPO",
    "LIST_ITEMS_UNIDAD_MEDIDA",
    "LIST_ITEMS_ESTADO",
    "LIST_MOVIMIENTOS_TIPO",
    "LIST_CLIENTES_TIPO",
    "LIST_CLIENTES_CANAL",
    "LIST_VENTAS_ESTADO_PAGO",
    "LIST_PAGOS_METODO",
    "LIST_PRODUCTOS_ESTADO"
]


def test_file_exists():
    """Test 1: Verify the Excel file exists"""
    print("\nTest 1: File Existence")
    print("-" * 50)

    if not OUTPUT_FILE.exists():
        print(f"[X] FAIL: File does not exist: {OUTPUT_FILE}")
        print(f"   Expected path: {OUTPUT_FILE.absolute()}")
        return False

    print(f"[OK] PASS: File exists")
    print(f"   Path: {OUTPUT_FILE.absolute()}")
    print(f"   Size: {OUTPUT_FILE.stat().st_size} bytes")
    return True


def test_sheets_count():
    """Test 2: Verify all expected sheets are present"""
    print("\nTest 2: Sheets Count")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_sheets = wb.sheetnames
        wb.close()

        # Check that all expected sheets are present (allow additional sheets)
        missing_sheets = []
        for expected_sheet in EXPECTED_SHEETS:
            if expected_sheet not in actual_sheets:
                missing_sheets.append(expected_sheet)

        if missing_sheets:
            print(f"[X] FAIL: Missing expected sheets")
            print(f"   Missing: {missing_sheets}")
            print(f"   Expected: {EXPECTED_SHEETS}")
            print(f"   Actual: {actual_sheets}")
            return False

        print(f"[OK] PASS: All expected sheets present ({len(EXPECTED_SHEETS)})")
        print(f"   Total sheets: {len(actual_sheets)}")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error loading workbook: {e}")
        return False


def test_sheets_order():
    """Test 3: Verify sheets are in correct order"""
    print("\nTest 3: Sheets Order")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_sheets = wb.sheetnames
        wb.close()

        mismatches = []
        for i, (expected, actual) in enumerate(zip(EXPECTED_SHEETS, actual_sheets)):
            if expected != actual:
                mismatches.append(f"Position {i}: expected '{expected}', got '{actual}'")

        if mismatches:
            print(f"[X] FAIL: Sheets not in correct order")
            for mismatch in mismatches:
                print(f"   {mismatch}")
            return False

        print(f"[OK] PASS: All sheets in correct order")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking sheets order: {e}")
        return False


def test_headers():
    """Test 4: Verify all expected headers are present"""
    print("\nTest 4: Headers Verification")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        all_pass = True

        for sheet_name, expected_headers in EXPECTED_HEADERS.items():
            if sheet_name not in wb.sheetnames:
                print(f"[X] FAIL: Sheet '{sheet_name}' not found")
                all_pass = False
                continue

            ws = wb[sheet_name]

            # Read headers from row 1
            actual_headers = [cell.value for cell in ws[1] if cell.value is not None]

            # Check that all expected headers are present (allow additional columns)
            missing_headers = []
            for expected_header in expected_headers:
                if expected_header not in actual_headers:
                    missing_headers.append(expected_header)

            if missing_headers:
                print(f"[X] FAIL: Missing headers in '{sheet_name}'")
                print(f"   Missing: {missing_headers}")
                print(f"   Expected: {expected_headers}")
                print(f"   Actual: {actual_headers}")
                all_pass = False
            else:
                print(f"[OK] PASS: All expected headers present in '{sheet_name}'")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking headers: {e}")
        return False


def test_defined_names():
    """Test 5: Verify defined names for validation lists"""
    print("\nTest 5: Defined Names (Validation Lists)")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_names = list(wb.defined_names.keys())
        wb.close()

        missing_names = []
        for expected_name in EXPECTED_DEFINED_NAMES:
            if expected_name not in actual_names:
                missing_names.append(expected_name)

        if missing_names:
            print(f"[X] FAIL: Missing defined names")
            for name in missing_names:
                print(f"   - {name}")
            print(f"\n   Expected: {EXPECTED_DEFINED_NAMES}")
            print(f"   Actual: {actual_names}")
            return False

        print(f"[OK] PASS: All defined names present ({len(EXPECTED_DEFINED_NAMES)})")
        print(f"   Names: {', '.join(EXPECTED_DEFINED_NAMES)}")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking defined names: {e}")
        return False


def test_data_validations():
    """Test 6: Verify at least one data validation exists"""
    print("\nTest 6: Data Validation")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        validation_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, 'data_validations'):
                validation_count += len(ws.data_validations)

        wb.close()

        if validation_count == 0:
            print(f"[X] FAIL: No data validations found")
            return False

        print(f"[OK] PASS: Data validations found ({validation_count} total)")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking data validations: {e}")
        return False


def test_tables():
    """Test 7: Verify Excel tables exist"""
    print("\nTest 7: Excel Tables")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        table_count = 0
        tables_info = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, 'tables'):
                for table_name, table_obj in ws.tables.items():
                    table_count += 1
                    tables_info.append(f"{sheet_name}: {table_name}")

        wb.close()

        if table_count == 0:
            print(f"[X] FAIL: No Excel tables found")
            return False

        print(f"[OK] PASS: Excel tables found ({table_count} total)")
        for info in tables_info:
            print(f"   - {info}")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking tables: {e}")
        return False

        print(f"[OK] PASS: Excel tables found ({table_count} total)")
        for info in tables_info:
            print(f"   - {info}")
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking tables: {e}")
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel MVP Skeleton Tests - Step 31")
    print("=" * 60)

    # Run all tests
    results = []
    results.append(("File Exists", test_file_exists()))
    results.append(("Sheets Count", test_sheets_count()))
    results.append(("Sheets Order", test_sheets_order()))
    results.append(("Headers", test_headers()))
    results.append(("Defined Names", test_defined_names()))
    results.append(("Data Validation", test_data_validations()))
    results.append(("Excel Tables", test_tables()))

    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] PASS" if result else "[X] FAIL"
        print(f"{status}: {test_name}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n*** All tests passed! ***")
        return 0
    else:
        print(f"\n*** {total - passed} test(s) failed ***")
        return 1


if __name__ == "__main__":
    exit(main())
