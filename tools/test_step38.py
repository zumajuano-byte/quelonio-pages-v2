"""
Tests for Excel MVP Step 38 - CSV Pack Round-trip (export + reimport) with FK validation

Tests verify:
- Export pack creates timestamped folder with CSV files
- Import validates FKs and prevents invalid data
- Row counts match between CSV and Excel
- FK relationships are 100% valid after round-trip
- Dashboard and auxiliary sheets remain intact
- KPI cells and formulas are preserved
"""

import sys
import os
import shutil
import tempfile
import uuid
from pathlib import Path

# Adjust paths for running from tools directory
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data" / "excel"
EXPORT_DIR = DATA_DIR / "exports"


def _repo_tmp_root():
    root = Path(os.getenv("ACI_TMP_DIR", "data/tmp"))
    root.mkdir(parents=True, exist_ok=True)
    return root


def _mkdtemp_repo(prefix):
    path = _repo_tmp_root() / f"{prefix}{uuid.uuid4().hex}"
    path.mkdir(parents=True, exist_ok=False)
    return path


def _latest_export_folder():
    if not EXPORT_DIR.exists():
        return None
    folders = [f for f in EXPORT_DIR.iterdir() if f.is_dir()]
    if not folders:
        return None
    return max(folders, key=lambda p: p.stat().st_mtime)


def test_round_trip_export():
    """Test that export pack creates timestamped folder"""
    print("\n[Test] Verifying round-trip export creates timestamped folder...")

    import subprocess
    result = subprocess.run([sys.executable, SCRIPT_DIR / "export_excel_csv_pack.py"], capture_output=True, text=True)

    if result.returncode != 0:
        print(f"  [FAIL] Export failed with return code {result.returncode}")
        print(f"  Output: {result.stdout}")
        print(f"  Error: {result.stderr}")
        return False

    # Check if export folder was created
    export_dir = Path("data/excel/exports")
    if not export_dir.exists():
        print("  [FAIL] Export directory not found")
        return False

    # Find most recent folder
    if not EXPORT_DIR.exists():
        print("  [FAIL] Export directory not found")
        return False

    # Find most recent folder
    latest_folder = _latest_export_folder()
    if latest_folder is None:
        print("  [FAIL] No export folders found")
        return False

    # Create backup of current Excel file
    excel_file = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")
    backup_file = excel_file.parent / f"backup_step38_{excel_file.stem}.xlsx"
    shutil.copy2(excel_file, backup_file)

    try:
        # Run import
        result = subprocess.run([sys.executable, "tools/import_excel_csv_pack.py", "--input", str(latest_folder), "--workbook", "data/excel/Quelonio_Excel_MVP_Skeleton.xlsx", "--dry-run", "0"],
                               capture_output=True, text=True)

        if result.returncode != 0:
            print(f"  [FAIL] Import failed with return code {result.returncode}")
            print(f"  Output: {result.stdout}")
            print(f"  Error: {result.stderr}")
            return False

        print(f"  [OK] Import completed successfully")
        return True

    finally:
        # Restore backup
        shutil.copy2(backup_file, excel_file)
        backup_file.unlink()


def test_import_with_fk_validation():
    """Test that import with FK validation succeeds with valid data"""
    print("\n[Test] Verifying import with FK validation succeeds...")

    # Find latest export folder
    export_dir = Path("data/excel/exports")
    latest_folder = _latest_export_folder()
    if latest_folder is None:
        print("  [FAIL] No export folders found")
        return False

    # Create backup of current Excel file
    excel_file = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")
    backup_file = excel_file.parent / f"backup_step38_{excel_file.stem}.xlsx"
    shutil.copy2(excel_file, backup_file)

    try:
        # Run import
        import subprocess
        result = subprocess.run([sys.executable, "tools/import_excel_csv_pack.py", "--input", str(latest_folder), "--workbook", "data/excel/Quelonio_Excel_MVP_Skeleton.xlsx", "--dry-run", "0"],
                               capture_output=True, text=True)

        if result.returncode != 0:
            print(f"  [FAIL] Import failed with return code {result.returncode}")
            print(f"  Output: {result.stdout}")
            print(f"  Error: {result.stderr}")
            return False

        print("  [OK] Import completed successfully")
        return True

    finally:
        # Restore backup
        shutil.copy2(backup_file, excel_file)
        backup_file.unlink()


def test_row_counts_match():
    """Test that row counts match between CSV and Excel after round-trip"""
    print("\n[Test] Verifying row counts match between CSV and Excel...")

    import openpyxl
    from import_excel_csv_pack import detect_delimiter, read_csv_file

    excel_file = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        print(f"  [FAIL] Could not load workbook: {e}")
        return False

    # Find latest export folder
    export_dir = Path("data/excel/exports")
    latest_folder = _latest_export_folder()
    if latest_folder is None:
        print("  [FAIL] No export folders found")
        return False

    # Check row counts for each sheet
    sheet_id_columns = {
        "01_Recetas": "receta_id",
        "02_RecetaVersiones": "recipe_version_id",
        "03_Lotes": "batch_id",
        "04_LoteMediciones": "medicion_id",
        "05_ItemsInventario": "item_id",
        "06_MovimientosInventario": "movimiento_id",
        "07_Proveedores": "proveedor_id",
        "08_LotesInsumo": "lote_insumo_id",
        "09_ConsumosLote": "consumo_id",
        "10_Productos": "producto_id",
        "11_Clientes": "cliente_id",
        "12_Ventas": "venta_id",
        "13_VentasLineas": "linea_id",
        "14_Pagos": "pago_id",
        "15_FulfillmentVentaLote": "fulfillment_id"
    }

    def get_excel_row_count(ws, id_column):
        """Get row count from Excel sheet"""
        header_map = {}
        for col_idx, cell in enumerate(ws[1], start=0):
            if cell.value:
                header_map[cell.value] = col_idx

        if id_column not in header_map:
            return 0

        col_idx = header_map[id_column]
        count = 0

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx + 1).value
            if cell_value is not None and str(cell_value).strip() != '':
                count += 1

        return count

    all_match = True
    for sheet_name in sheet_id_columns:
        if sheet_name not in wb.sheetnames:
            continue

        # Get Excel row count
        ws = wb[sheet_name]
        id_column = sheet_id_columns[sheet_name]
        excel_count = get_excel_row_count(ws, id_column)

        # Get CSV row count
        csv_file = latest_folder / f"{sheet_name}.csv"
        if csv_file.exists():
            delimiter = detect_delimiter(csv_file)
            headers, data_rows = read_csv_file(csv_file, delimiter)
            csv_count = len([row for row in data_rows if any(cell is not None and str(cell).strip() != '' for cell in row)])
        else:
            csv_count = 0

        if excel_count != csv_count:
            print(f"  [FAIL] {sheet_name}: Excel has {excel_count} rows, CSV has {csv_count} rows")
            all_match = False
        else:
            print(f"  [OK] {sheet_name}: {excel_count} rows match")

    wb.close()

    if all_match:
        print("  [OK] All row counts match")
        return True
    else:
        print("  [FAIL] Some row counts don't match")
        return False


def test_fk_validation_prevents_invalid_data():
    """Test that FK validation prevents import of invalid data"""
    print("\n[Test] Verifying FK validation prevents invalid data...")

    import subprocess
    from import_excel_csv_pack import detect_delimiter, read_csv_file

    # Find latest export folder
    export_dir = Path("data/excel/exports")
    latest_folder = _latest_export_folder()
    if latest_folder is None:
        print("  [FAIL] No export folders found")
        return False

    temp_path = _mkdtemp_repo("step38_fk_")
    try:
        # Copy all pack files (including manifest) so altered CSV is detected deterministically.
        for src_file in latest_folder.iterdir():
            if src_file.is_file():
                shutil.copy2(src_file, temp_path / src_file.name)

        # Modify one CSV to have invalid FK
        invalid_csv = temp_path / "02_RecetaVersiones.csv"
        if invalid_csv.exists():
            # Read original
            delimiter = detect_delimiter(invalid_csv)
            headers, data_rows = read_csv_file(invalid_csv, delimiter)

            if data_rows:
                # Make first row have invalid receta_id
                header_map = {h: i for i, h in enumerate(headers)}
                if "receta_id" in header_map:
                    col_idx = header_map["receta_id"]
                    data_rows[0][col_idx] = "INVALID-REC-ID"

                    # Write back
                    import csv
                    with open(invalid_csv, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f, delimiter=delimiter, quotechar='"')
                        writer.writerow(headers)
                        writer.writerows(data_rows)

        # Try to import - should fail
        result = subprocess.run([sys.executable, "tools/import_excel_csv_pack.py", "--input", str(temp_path), "--workbook", "data/excel/Quelonio_Excel_MVP_Skeleton.xlsx", "--dry-run", "0"],
                               capture_output=True, text=True)

        if result.returncode == 0:
            print("  [FAIL] Import succeeded when it should have failed due to FK violation")
            return False
        else:
            print("  [OK] Import correctly failed due to FK violation")
            return True
    finally:
        shutil.rmtree(temp_path, ignore_errors=True)


def test_dashboard_sheets_preserved():
    """Test that dashboard and auxiliary sheets remain intact (if present)"""
    print("\n[Test] Verifying dashboard and auxiliary sheets preserved...")

    import openpyxl

    excel_file = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

    try:
        wb = openpyxl.load_workbook(excel_file)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        print(f"  [FAIL] Could not load workbook: {e}")
        return False

    # Check for required sheets that should always be present
    required_sheets = ["99_Listas"]
    missing_required = [s for s in required_sheets if s not in sheets]
    if missing_required:
        print(f"  [FAIL] Missing required sheets: {missing_required}")
        return False

    # Check for optional dashboard sheets
    optional_sheets = ["20_Dashboard", "21_Calc"]
    missing_optional = [s for s in optional_sheets if s not in sheets]
    if missing_optional:
        print(f"  [INFO] Optional dashboard sheets not present: {missing_optional} (expected in later steps)")
    else:
        print("  [OK] Dashboard and auxiliary sheets present")

    return True


def test_kpi_cells_preserved():
    """Test that KPI cells and formulas are preserved (if dashboard exists)"""
    print("\n[Test] Verifying KPI cells and formulas preserved...")

    import openpyxl

    excel_file = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=False)
        if "20_Dashboard" not in wb.sheetnames:
            wb.close()
            print("  [INFO] Dashboard sheet not present (expected in later steps)")
            return True
        ws = wb["20_Dashboard"]
    except Exception as e:
        print(f"  [FAIL] Could not load workbook: {e}")
        return False

    # Check specific KPI cells
    kpi_checks = [
        (6, 3, "KPI_Ventas_Periodo"),  # Row 6, Col 3
        (8, 3, "KPI_SaldoPorCobrar"),  # Row 8, Col 3
    ]

    all_ok = True
    for row, col, kpi_name in kpi_checks:
        formula = ws.cell(row=row, column=col).value
        if formula and isinstance(formula, str) and formula.startswith("="):
            print(f"  [OK] {kpi_name} has formula")
        else:
            print(f"  [FAIL] {kpi_name} missing formula: {formula}")
            all_ok = False

    wb.close()

    if all_ok:
        print("  [OK] All KPI formulas preserved")
        return True
    else:
        print("  [FAIL] Some KPI formulas missing")
        return False


def run_all_tests():
    """Run all Step 38 tests."""
    print("=" * 60)
    print("Excel MVP Step 38 Tests - CSV Pack Round-trip with FK validation")
    print("=" * 60)

    test1 = test_round_trip_export()
    test2 = test_import_with_fk_validation()
    test3 = test_row_counts_match()
    test4 = test_fk_validation_prevents_invalid_data()
    test5 = test_dashboard_sheets_preserved()
    test6 = test_kpi_cells_preserved()

    print("\n" + "=" * 60)
    print("Test Summary")
    print("=" * 60)
    print(f"Round-trip export:                {'PASS' if test1 else 'FAIL'}")
    print(f"Import with FK validation:        {'PASS' if test2 else 'FAIL'}")
    print(f"Row counts match:                 {'PASS' if test3 else 'FAIL'}")
    print(f"FK validation prevents invalid:   {'PASS' if test4 else 'FAIL'}")
    print(f"Dashboard sheets preserved:       {'PASS' if test5 else 'FAIL'}")
    print(f"KPI cells preserved:              {'PASS' if test6 else 'FAIL'}")
    print("=" * 60)

    all_passed = all([test1, test2, test3, test4, test5, test6])

    if all_passed:
        print("[OK] ALL TESTS PASSED")
        print("*** All tests passed! ***")
        return 0
    else:
        print("[FAIL] SOME TESTS FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(run_all_tests())
