"""Import CSV pack to Excel with validations

Imports CSV files from pack folder to Excel, validates manifest,
checks uniqueness, enums, types, and preserves non-table sheets.

Usage:
    python tools/import_excel_csv_pack.py --input <folder> --workbook <excel> [--dry-run 0|1]
"""

import argparse
import csv
import hashlib
import json
import openpyxl
import os
import shutil
import sys
from datetime import datetime


# Default enum values (if no catalogs)
DEFAULT_ENUMS = {
    "03_Lotes.estado": ["planificado", "coccion", "fermentacion", "maduracion", "envasado", "completado", "cancelado"],
    "06_MovimientosInventario.tipo_movimiento": [],  # Will be loaded from CAT_MovTipo
    "05_ItemsInventario.tipo": [],  # Will be loaded from CAT_ItemTipo
    "11_Clientes.tipo": [],  # Will be loaded from CAT_ClienteTipo
    "11_Clientes.canal": [],  # Will be loaded from CAT_Canal
    "12_Ventas.estado_pago": []  # Will be loaded from CAT_PagoEstado
}

# ID columns for uniqueness check
ID_COLUMNS = {
    "01_Recetas": "receta_id",
    "02_RecetaVersiones": "recipe_version_id",
    "03_Lotes": "batch_id",
    "04_LoteMediciones": "medicion_id",
    "05_ItemsInventario": "item_id",
    "06_MovimientosInventario": "movimiento_id",
    "07_Proveedores": "proveedor_id",
    "08_LotesInsumo": "lote_insumo_id",
    "09_ConsumosLote": "consumo_id",
    "10_Productos": "producto_id",
    "11_Clientes": "cliente_id",
    "12_Ventas": "venta_id",
    "13_VentasLineas": "linea_id",
    "14_Pagos": "pago_id",
    "15_FulfillmentVentaLote": "fulfillment_id"
}


def load_manifest(pack_folder):
    """Load and validate manifest."""
    manifest_path = os.path.join(pack_folder, 'pack_manifest.json')
    if not os.path.exists(manifest_path):
        return None

    try:
        with open(manifest_path, 'r', encoding='utf-8') as f:
            manifest = json.load(f)
        return manifest
    except Exception as e:
        print(f"[ERROR] Failed to load manifest: {e}")
        return None


def validate_manifest(manifest):
    """Validate manifest structure."""
    required_keys = ['schema_version', 'total_files', 'files']
    for key in required_keys:
        if key not in manifest:
            print(f"[ERROR] Manifest missing key: {key}")
            return False

    if manifest['schema_version'] != 'mvp-v1':
        print(f"[ERROR] Invalid schema_version: {manifest['schema_version']}")
        return False

    if manifest['total_files'] != len(manifest['files']):
        print(f"[ERROR] total_files mismatch: {manifest['total_files']} vs {len(manifest['files'])}")
        return False

    return True


def check_csv_files(pack_folder, manifest):
    """Check that all CSV files exist and match manifest."""
    for file_info in manifest['files']:
        csv_path = os.path.join(pack_folder, file_info['name'])
        if not os.path.exists(csv_path):
            print(f"[ERROR] Missing CSV file: {file_info['name']}")
            return False

        # Check SHA256
        actual_sha256 = calculate_sha256(csv_path)
        if actual_sha256 != file_info['sha256']:
            print(f"[ERROR] SHA256 mismatch for {file_info['name']}: expected {file_info['sha256']}, got {actual_sha256}")
            return False

    return True


def calculate_sha256(file_path):
    """Calculate SHA256 of file bytes."""
    hash_sha256 = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()


def load_catalogs(wb):
    """Load enum values from Catalogos or 99_Listas sheet."""
    catalogs = {}

    # Try Catalogos first, then 99_Listas
    catalog_sheet = None
    if "Catalogos" in wb.sheetnames:
        catalog_sheet = wb["Catalogos"]
    elif "99_Listas" in wb.sheetnames:
        catalog_sheet = wb["99_Listas"]

    if catalog_sheet:
        # Read catalog data
        for row in range(2, catalog_sheet.max_row + 1):
            cat_name = catalog_sheet.cell(row=row, column=1).value
            valor = catalog_sheet.cell(row=row, column=2).value
            if cat_name and valor is not None:
                if cat_name not in catalogs:
                    catalogs[cat_name] = []
                catalogs[cat_name].append(str(valor))

    return catalogs


def get_enum_values(sheet_name, column_name, catalogs):
    """Get enum values for a column from catalogs or defaults."""
    key = f"{sheet_name}.{column_name}"
    if key in DEFAULT_ENUMS and DEFAULT_ENUMS[key]:
        return DEFAULT_ENUMS[key]

    # Map to catalog names
    catalog_map = {
        "03_Lotes.estado": "CAT_LoteEstado",
        "06_MovimientosInventario.tipo_movimiento": "CAT_MovTipo",
        "05_ItemsInventario.tipo": "CAT_ItemTipo",
        "11_Clientes.tipo": "CAT_ClienteTipo",
        "11_Clientes.canal": "CAT_Canal",
        "12_Ventas.estado_pago": "CAT_PagoEstado"
    }

    if key in catalog_map:
        cat_name = catalog_map[key]
        if cat_name in catalogs:
            return catalogs[cat_name]

    return None  # No validation if not found


def validate_csv_data(csv_path, sheet_name, delimiter, bom, catalogs):
    """Validate CSV data: uniqueness, enums, types."""
    errors = []

    # Read CSV
    data = []
    encoding = 'utf-8-sig' if bom else 'utf-8'
    try:
        with open(csv_path, 'r', encoding=encoding) as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            data = list(reader)
    except Exception as e:
        errors.append(f"Failed to read CSV: {e}")
        return errors

    if not data:
        return errors  # Empty CSV is OK

    # Check ID uniqueness
    id_column = ID_COLUMNS.get(sheet_name)
    if id_column and id_column in data[0]:
        ids = []
        for row_idx, row in enumerate(data, 1):
            id_val = row.get(id_column)
            if id_val is not None and id_val != '':
                if id_val in ids:
                    errors.append(f"Duplicate {id_column} '{id_val}' in row {row_idx}")
                ids.append(id_val)

    # Validate enums
    enum_columns = {
        "03_Lotes": ["estado"],
        "06_MovimientosInventario": ["tipo_movimiento"],
        "05_ItemsInventario": ["tipo"],
        "11_Clientes": ["tipo", "canal"],
        "12_Ventas": ["estado_pago"]
    }

    if sheet_name in enum_columns:
        for col in enum_columns[sheet_name]:
            if col in data[0]:
                enum_values = get_enum_values(sheet_name, col, catalogs)
                if enum_values:
                    for row_idx, row in enumerate(data, 1):
                        val = row.get(col)
                        if val and val not in enum_values:
                            errors.append(f"Invalid {col} '{val}' in row {row_idx}, expected: {enum_values}")

    # Validate types
    date_columns = [col for col in data[0].keys() if '_fecha' in col.lower() or col.lower().endswith('fecha') or col.lower() in ['fecha_inicio', 'fecha_fin', 'fecha_creacion', 'fecha_actualizacion']]
    numeric_columns = ['volumen_litros', 'og_target', 'fg_target', 'cantidad', 'precio_unitario', 'total', 'stock_actual', 'stock_minimo']

    for row_idx, row in enumerate(data, 1):
        for col, val in row.items():
            if not val:  # Empty is OK
                continue

            # Date validation
            if col in date_columns:
                try:
                    datetime.strptime(val, '%Y-%m-%d')
                except ValueError:
                    errors.append(f"Invalid date '{val}' in {col} row {row_idx}, expected YYYY-MM-DD")

            # Numeric validation
            elif col in numeric_columns:
                try:
                    float(val.replace(',', '.'))  # Handle decimal comma
                except ValueError:
                    errors.append(f"Invalid number '{val}' in {col} row {row_idx}")

    return errors


def import_csv_to_sheet(wb, csv_path, sheet_name, delimiter, bom):
    """Import CSV data to worksheet."""
    # Create or clear sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Clear existing data
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    # Read CSV and write to sheet
    encoding = 'utf-8-sig' if bom else 'utf-8'
    with open(csv_path, 'r', encoding=encoding) as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, cell_value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)


def import_csv_pack(pack_folder, workbook_path, dry_run=False):
    """Main import function."""
    print("Importing Excel CSV Pack...")
    print(f"Pack folder: {pack_folder}")
    print(f"Workbook: {workbook_path}")
    print(f"Dry run: {dry_run}")

    # Validate inputs
    if not os.path.exists(pack_folder):
        print(f"[ERROR] Pack folder not found: {pack_folder}")
        return False

    if not os.path.exists(workbook_path):
        print(f"[ERROR] Workbook not found: {workbook_path}")
        return False

    # Load manifest
    manifest = load_manifest(pack_folder)
    if manifest:
        print("Manifest found, validating...")
        if not validate_manifest(manifest):
            return False
        if not check_csv_files(pack_folder, manifest):
            return False
        delimiter = get_delimiter(manifest['delimiter'])
        bom = manifest['bom']
        print(f"Manifest OK: {manifest['total_files']} files, delimiter={manifest['delimiter']}, bom={bom}")
    else:
        print("[WARNING] No manifest found, using defaults")
        delimiter = ';'
        bom = False

    # Load workbook
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as e:
        print(f"[ERROR] Failed to load workbook: {e}")
        return False

    # Load catalogs for enum validation
    catalogs = load_catalogs(wb)
    print(f"Loaded catalogs: {list(catalogs.keys())}")

    # Validate each CSV
    validation_errors = []
    for file_info in manifest['files'] if manifest else []:
        csv_path = os.path.join(pack_folder, file_info['name'])
        sheet_name = file_info['name'].replace('.csv', '')
        print(f"Validating {file_info['name']}...")

        errors = validate_csv_data(csv_path, sheet_name, delimiter, bom, catalogs)
        if errors:
            validation_errors.extend([f"{file_info['name']}: {err}" for err in errors])

    if validation_errors:
        print("[ERROR] Validation failed:")
        for err in validation_errors[:10]:  # Show first 10
            print(f"  {err}")
        if len(validation_errors) > 10:
            print(f"  ... and {len(validation_errors) - 10} more errors")
        return False

    print("All validations passed!")

    if dry_run:
        wb.close()
        print("[OK] Dry run complete - no changes saved")
        return True

    # Import data
    for file_info in manifest['files'] if manifest else []:
        csv_path = os.path.join(pack_folder, file_info['name'])
        sheet_name = file_info['name'].replace('.csv', '')
        print(f"Importing {file_info['name']} to {sheet_name}...")

        try:
            import_csv_to_sheet(wb, csv_path, sheet_name, delimiter, bom)
            print(f"  Imported {file_info['rows']} rows")
        except Exception as e:
            print(f"[ERROR] Failed to import {file_info['name']}: {e}")
            wb.close()
            return False

    # Save workbook
    try:
        wb.save(workbook_path)
        print(f"[OK] Workbook saved: {workbook_path}")
    except Exception as e:
        print(f"[ERROR] Failed to save workbook: {e}")
        wb.close()
        return False

    wb.close()
    print("[OK] Import complete!")
    return True


def get_delimiter(delimiter_str):
    """Convert delimiter string to character."""
    if delimiter_str == 'comma':
        return ','
    elif delimiter_str == 'semicolon':
        return ';'
    elif delimiter_str == 'tab':
        return '\t'
    else:
        raise ValueError(f"Invalid delimiter: {delimiter_str}")


def detect_delimiter(csv_path):
    """Auto-detect delimiter from a CSV file path."""
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[',', ';', '\t'])
        return dialect.delimiter
    except Exception:
        if sample.count(';') >= sample.count(',') and sample.count(';') >= sample.count('\t'):
            return ';'
        if sample.count('\t') > sample.count(','):
            return '\t'
        return ','


def read_csv_file(csv_path, delimiter):
    """Read a CSV file and return (headers, data_rows) for compatibility tests."""
    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def main():
    parser = argparse.ArgumentParser(description="Import CSV pack to Excel with validations")
    parser.add_argument('--input', required=True, help='Pack folder path')
    parser.add_argument('--workbook', required=True, help='Excel workbook path')
    parser.add_argument('--dry-run', type=int, choices=[0, 1], default=0, help='Dry run mode (default: 0)')

    args = parser.parse_args()

    success = import_csv_pack(args.input, args.workbook, bool(args.dry_run))
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
