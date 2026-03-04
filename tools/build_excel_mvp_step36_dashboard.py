"""Build Excel MVP with Dashboard (Step 36)

Creates/updates the Excel MVP with operational dashboard, KPIs, and mini-panels.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os
import shutil

# Constants
EXCEL_PATH = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
BACKUP_DIR = os.path.join("data", "excel", "_backups")

# KPI Configuration
KPIS = [
    {"name": "KPI_LotesEnCurso", "label": "Lotes en Curso", "row": 2, "col": 2},
    {"name": "KPI_StockCriticoCount", "label": "Items Stock Crítico", "row": 4, "col": 2},
    {"name": "KPI_Ventas_30d", "label": "Ventas (últimos 30 días)", "row": 6, "col": 2},
    {"name": "KPI_SaldoPorCobrar", "label": "Saldo por Cobrar", "row": 8, "col": 2},
    {"name": "KPI_MargenEstimado", "label": "Margen Estimado (%)", "row": 10, "col": 2},
]

# Mini-panel configurations
PANEL_LOTES_START = {"row": 13, "col": 1}
PANEL_STOCK_START = {"row": 13, "col": 6}
PANEL_VENTAS_START = {"row": 24, "col": 1}


def create_backup():
    """Create backup of existing Excel file with timestamp."""
    if not os.path.exists(EXCEL_PATH):
        return None
    
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"Quelonio_Excel_MVP_Skeleton_step36_{timestamp}.xlsx")
    shutil.copy2(EXCEL_PATH, backup_path)
    print(f"[OK] Backup created: {backup_path}")
    return backup_path


def init_workbook():
    """Initialize workbook and create base sheets."""
    wb = None
    
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    
    return wb


def create_table(ws, table_name, start_row, start_col, headers):
    """Create Excel table with structured references."""
    # Add headers
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Set column widths
    for col_idx, header in enumerate(headers):
        ws.column_dimensions[get_column_letter(start_col + col_idx)].width = 15
    
    return start_row, start_col, len(headers), 1


def add_03_Lotes_sheet(wb):
    """Create or update 03_Lotes sheet with sample data."""
    if "03_Lotes" in wb.sheetnames:
        ws = wb["03_Lotes"]
    else:
        ws = wb.create_sheet("03_Lotes", 0)
    
    headers = ["id", "nombre", "producto", "estado", "fecha_inicio", "fecha_fin", "cantidad"]
    start_row, start_col, num_cols, num_rows = create_table(ws, "TblLotes", 1, 1, headers)
    
    # Sample data for lots in various states
    estados_en_curso = ["coccion", "fermentacion", "maduracion", "envasado", "planificado"]
    sample_data = [
        ["L001", "Lote Queso 1", "Queso Madurado", "coccion", datetime(2026, 1, 10), datetime(2026, 1, 20), 150],
        ["L002", "Lote Queso 2", "Queso Fresco", "fermentacion", datetime(2026, 1, 12), datetime(2026, 1, 22), 200],
        ["L003", "Lote Queso 3", "Queso Semi", "maduracion", datetime(2026, 1, 8), datetime(2026, 1, 25), 120],
        ["L004", "Lote Queso 4", "Queso Azul", "envasado", datetime(2026, 1, 14), datetime(2026, 1, 28), 80],
        ["L005", "Lote Queso 5", "Queso Crema", "planificado", datetime(2026, 1, 15), datetime(2026, 1, 30), 180],
        ["L006", "Lote Queso 6", "Queso Ahumado", "completado", datetime(2026, 1, 5), datetime(2026, 1, 10), 100],
        ["L007", "Lote Queso 7", "Queso Brie", "coccion", datetime(2026, 1, 13), datetime(2026, 1, 23), 90],
    ]
    
    for row_idx, row_data in enumerate(sample_data):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=value)
    
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    ws.title = "03_Lotes"
    print("[OK] 03_Lotes sheet created/updated")


def add_05_ItemsInventario_sheet(wb):
    """Create or update 05_ItemsInventario sheet with stock_minimo column."""
    if "05_ItemsInventario" in wb.sheetnames:
        ws = wb["05_ItemsInventario"]
        
        # Check if stock_minimo exists, add if not
        header_row = 1
        existing_headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header:
                existing_headers.append(header)
        
        if "stock_minimo" not in existing_headers:
            # Add stock_minimo column
            last_col = ws.max_column
            new_col = last_col + 1
            ws.cell(row=header_row, column=new_col, value="stock_minimo")
            ws.cell(row=header_row, column=new_col).font = Font(bold=True)
            ws.cell(row=header_row, column=new_col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            ws.cell(row=header_row, column=new_col).alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(new_col)].width = 15
            
            # Fill default values for existing rows
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=new_col, value=0)
            
            print("[OK] Added stock_minimo column to 05_ItemsInventario")
    else:
        ws = wb.create_sheet("05_ItemsInventario", 1)
        headers = ["id", "nombre", "categoria", "stock_actual", "stock_minimo", "unidad", "precio_unitario"]
        start_row, start_col, num_cols, num_rows = create_table(ws, "TblItems", 1, 1, headers)
        
        sample_data = [
            ["I001", "Queso Madurado", "Lácteos", 50, 100, "kg", 2500],
            ["I002", "Queso Fresco", "Lácteos", 200, 150, "kg", 1800],
            ["I003", "Queso Semi", "Lácteos", 30, 50, "kg", 2100],
            ["I004", "Queso Azul", "Lácteos", 45, 40, "kg", 3200],
            ["I005", "Queso Crema", "Lácteos", 80, 100, "kg", 2800],
            ["I006", "Queso Ahumado", "Lácteos", 60, 70, "kg", 3000],
            ["I007", "Queso Brie", "Lácteos", 25, 30, "kg", 3500],
        ]
        
        for row_idx, row_data in enumerate(sample_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=value)
        
        ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
        print("[OK] 05_ItemsInventario sheet created with stock_minimo")
    
    ws.title = "05_ItemsInventario"


def add_11_Clientes_sheet(wb):
    """Create or update 11_Clientes sheet."""
    if "11_Clientes" in wb.sheetnames:
        ws = wb["11_Clientes"]
    else:
        ws = wb.create_sheet("11_Clientes", 2)
    
    headers = ["id", "nombre", "canal", "region", "email", "telefono"]
    start_row, start_col, num_cols, num_rows = create_table(ws, "TblClientes", 1, 1, headers)
    
    sample_data = [
        ["C001", "Supermercado Central", "mayorista", "Norte", "central@super.com", "555-0101"],
        ["C002", "Tienda Vecinal", "minorista", "Centro", "tienda@vecinal.com", "555-0102"],
        ["C003", "Distribuidor Regional", "mayorista", "Sur", "distri@regional.com", "555-0103"],
        ["C004", "Restaurante Gourmet", "horeca", "Centro", "gourmet@rest.com", "555-0104"],
        ["C005", "Hotel Luxury", "horeca", "Norte", "luxury@hotel.com", "555-0105"],
    ]
    
    for row_idx, row_data in enumerate(sample_data):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=value)
    
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    ws.title = "11_Clientes"
    print("[OK] 11_Clientes sheet created/updated")


def add_12_Ventas_sheet(wb):
    """Create or update 12_Ventas sheet with calculated columns."""
    if "12_Ventas" in wb.sheetnames:
        ws = wb["12_Ventas"]
    else:
        ws = wb.create_sheet("12_Ventas", 3)
    
    headers = ["id", "cliente_id", "fecha", "cantidad", "precio_unitario", "total_calc", "pagado", "saldo_calc"]
    start_row, start_col, num_cols, num_rows = create_table(ws, "TblVentas", 1, 1, headers)
    
    # Generate sample data for last 45 days
    today = datetime.now()
    sample_data = []
    for i in range(20):
        days_ago = i * 2
        fecha = today - timedelta(days=days_ago)
        cliente_idx = i % 5
        cantidad = [10, 25, 15, 5, 8][i % 5]
        precio = [2500, 1800, 2100, 3200, 2800][i % 5]
        total = cantidad * precio
        pagado = total * [1.0, 0.8, 0.5, 1.0, 0.3][i % 5]
        saldo = total - pagado
        
        sample_data.append([
            f"V{str(i+1).zfill(3)}",
            f"C00{cliente_idx + 1}",
            fecha,
            cantidad,
            precio,
            total,
            pagado,
            saldo
        ])
    
    for row_idx, row_data in enumerate(sample_data):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=start_row + 1 + row_idx, column=start_col + col_idx, value=value)
    
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    ws.title = "12_Ventas"
    print("[OK] 12_Ventas sheet created/updated")


def add_21_Calc_sheet(wb):
    """Create 21_Calc sheet for auxiliary calculations."""
    if "21_Calc" in wb.sheetnames:
        wb.remove(wb["21_Calc"])
    
    ws = wb.create_sheet("21_Calc", len(wb.sheetnames))
    
    headers = ["Calculation", "Description", "Formula", "Result"]
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=col_idx + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 20
    
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.title = "21_Calc"
    print("[OK] 21_Calc sheet created")


def add_20_Dashboard_sheet(wb):
    """Create 20_Dashboard sheet with KPIs and mini-panels."""
    if "20_Dashboard" in wb.sheetnames:
        wb.remove(wb["20_Dashboard"])
    
    ws = wb.create_sheet("20_Dashboard", len(wb.sheetnames))
    
    # Title
    title_cell = ws.cell(row=1, column=1, value="Dashboard Operativo - Quelonio")
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # KPIs section title
    kpi_title = ws.cell(row=2, column=1, value="KPIs Principales")
    kpi_title.font = Font(size=12, bold=True, color="FFFFFF")
    kpi_title.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    kpi_title.alignment = Alignment(horizontal="center", vertical="center")
    
    for kpi in KPIS:
        # Label
        label_cell = ws.cell(row=kpi["row"], column=kpi["col"], value=kpi["label"])
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        
        # KPI name (below label)
        name_cell = ws.cell(row=kpi["row"] + 1, column=kpi["col"], value=kpi["name"])
        name_cell.font = Font(italic=True, size=9, color="666666")
        name_cell.alignment = Alignment(horizontal="right")

        # Preserve KPI identifiers on the KPI row (stable location for contract tests).
        stable_name_cell = ws.cell(row=kpi["row"], column=4, value=kpi["name"])
        stable_name_cell.font = Font(size=8, color="999999")
        stable_name_cell.alignment = Alignment(horizontal="left")
        
        # Formula cell (formatted nicely without merging)
        formula_cell = ws.cell(row=kpi["row"], column=kpi["col"] + 1, value="")
        formula_cell.font = Font(size=14, bold=True, color="4472C4")
        formula_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        formula_cell.alignment = Alignment(horizontal="center")
        formula_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add formulas for KPIs (note: row + 1 for the name cell)
    ws.cell(row=2, column=3, value='=COUNTIFS(03_Lotes!D:D,"coccion",03_Lotes!D:D,"fermentacion",03_Lotes!D:D,"maduracion",03_Lotes!D:D,"envasado",03_Lotes!D:D,"planificado")')
    ws.cell(row=4, column=3, value='=COUNTIFS(05_ItemsInventario!D:D,"<=0",05_ItemsInventario!E:E,"<=0")')
    ws.cell(row=6, column=3, value='=SUMIFS(12_Ventas!F:F,12_Ventas!C:C,">="&TODAY()-30)')
    ws.cell(row=8, column=3, value='=SUM(12_Ventas!H:H)')
    ws.cell(row=10, column=3, value='=AVERAGE(12_Ventas!F:F)*0.25')
    
    # Mini-panel: Lotes en curso
    ws.cell(row=13, column=1, value="Lotes en Curso").font = Font(bold=True, size=11)
    lote_headers = ["ID", "Nombre", "Producto", "Estado", "Fecha Inicio", "Cantidad"]
    for col_idx, header in enumerate(lote_headers):
        cell = ws.cell(row=14, column=1 + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Formula to extract lots in progress
    for row_idx in range(1, 8):
        ws.cell(row=14 + row_idx, column=1, value=f'=FILTER(03_Lotes!A2:F100, (03_Lotes!D2:D100="coccion")+(03_Lotes!D2:D100="fermentacion")+(03_Lotes!D2:D100="maduracion")+(03_Lotes!D2:D100="envasado")+(03_Lotes!D2:D100="planificado"),"")')
    
    # Mini-panel: Stock crítico
    ws.cell(row=13, column=6, value="Stock Crítico").font = Font(bold=True, size=11)
    stock_headers = ["ID", "Nombre", "Stock Actual", "Stock Mínimo", "Estado"]
    for col_idx, header in enumerate(stock_headers):
        cell = ws.cell(row=14, column=6 + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Formula to extract critical stock
    ws.cell(row=15, column=6, value='=IFERROR(FILTER(05_ItemsInventario!A2:E100, 05_ItemsInventario!D2:D100<=05_ItemsInventario!E2:E100,""),"Sin stock crítico")')
    
    # Mini-panel: Ventas y cobranzas
    ws.cell(row=24, column=1, value="Ventas y Cobranzas por Canal").font = Font(bold=True, size=11)
    ventas_headers = ["Canal", "Ventas Totales", "Saldo por Cobrar"]
    for col_idx, header in enumerate(ventas_headers):
        cell = ws.cell(row=25, column=1 + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data for sales by channel
    canales = ["mayorista", "minorista", "horeca"]
    for row_idx, canal in enumerate(canales):
        ws.cell(row=26 + row_idx, column=1, value=canal)
        ws.cell(row=26 + row_idx, column=2, value=f'=SUMIFS(12_Ventas!F:F,INDEX(11_Clientes!A:C,MATCH(12_Ventas!B2:B100,11_Clientes!A:A,0),2),"{canal}")')
        ws.cell(row=26 + row_idx, column=3, value=f'=SUMIFS(12_Ventas!H:H,INDEX(11_Clientes!A:C,MATCH(12_Ventas!B2:B100,11_Clientes!A:A,0),2),"{canal}")')
    
    # Format column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.title = "20_Dashboard"
    print("[OK] 20_Dashboard sheet created with KPIs and mini-panels")


def build():
    """Main build function."""
    print("Building Excel MVP - Step 36: Dashboard with KPIs")
    print("-" * 50)
    
    # Create backup
    create_backup()
    
    # Initialize workbook
    wb = init_workbook()
    
    # Preserve core Step31-35 sheets and only add dashboard/calc sheets.
    add_21_Calc_sheet(wb)
    add_20_Dashboard_sheet(wb)
    
    # Ensure data directory exists
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    # Save workbook
    wb.save(EXCEL_PATH)
    print("-" * 50)
    print(f"[OK] Excel MVP saved to: {EXCEL_PATH}")
    print("[OK] Step 36 build complete!")


if __name__ == "__main__":
    build()
