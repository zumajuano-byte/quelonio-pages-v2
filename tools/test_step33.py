#!/usr/bin/env python
"""
Test script for Excel MVP Step 33
Verifica: 00_Home, 98_Ayuda, Hyperlinks, Defined Names, Compatibilidad
"""

import os
from pathlib import Path
from openpyxl import load_workbook

# Configuration
OUTPUT_FILE = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

# Expected Step 32 catalog defined names
EXPECTED_STEP32_NAMES = [
    "lst_lote_estado",
    "lst_mov_tipo",
    "lst_pago_estado",
    "lst_cliente_tipo",
    "lst_canal",
    "lst_item_tipo",
    "lst_receta_id",
    "lst_recipe_version_id",
    "lst_batch_id",
    "lst_item_id",
    "lst_lote_insumo_id",
    "lst_venta_id",
    "lst_linea_id",
    "lst_producto_id"
]

# Expected navigation sheets (to be linked from Home)
EXPECTED_NAVIGATION_SHEETS = [
    "01_Recetas",
    "02_RecetaVersiones",
    "03_Lotes",
    "05_ItemsInventario",
    "08_LotesInsumo",
    "09_ConsumosLote",
    "10_Productos",
    "11_Clientes",
    "12_Ventas",
    "13_VentasLineas",
    "14_Pagos",
    "15_FulfillmentVentaLote"
]


def test_file_exists():
    """Test 1: Verify the Excel file exists"""
    print("\nTest 1: File Existence")
    print("-" * 50)

    if not OUTPUT_FILE.exists():
        print(f"[X] FAIL: File does not exist: {OUTPUT_FILE}")
        return False

    print(f"[OK] PASS: File exists")
    return True


def test_home_sheet_exists_and_first():
    """Test 2: Verify 00_Home exists and is first sheet"""
    print("\nTest 2: 00_Home Sheet Exists and is First")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        if "00_Home" not in wb.sheetnames:
            print(f"[X] FAIL: 00_Home sheet not found")
            print(f"   Available sheets: {wb.sheetnames}")
            wb.close()
            return False

        # Check it's the first sheet
        first_sheet = wb.sheetnames[0]
        if first_sheet != "00_Home":
            print(f"[X] FAIL: 00_Home is not the first sheet")
            print(f"   First sheet: {first_sheet}")
            wb.close()
            return False

        print(f"[OK] PASS: 00_Home exists and is first sheet")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking 00_Home sheet: {e}")
        return False


def test_ayuda_sheet_exists():
    """Test 3: Verify 98_Ayuda sheet exists"""
    print("\nTest 3: 98_Ayuda Sheet Existence")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        if "98_Ayuda" not in wb.sheetnames:
            print(f"[X] FAIL: 98_Ayuda sheet not found")
            print(f"   Available sheets: {wb.sheetnames}")
            wb.close()
            return False

        print(f"[OK] PASS: 98_Ayuda sheet exists")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking 98_Ayuda sheet: {e}")
        return False


def test_catalogos_sheet_exists():
    """Test 4: Verify Catalogos sheet exists (from Step 32)"""
    print("\nTest 4: Catalogos Sheet Existence")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        if "Catalogos" not in wb.sheetnames:
            print(f"[X] FAIL: Catalogos sheet not found")
            print(f"   Available sheets: {wb.sheetnames}")
            wb.close()
            return False

        print(f"[OK] PASS: Catalogos sheet exists")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking Catalogos sheet: {e}")
        return False


def test_step32_defined_names():
    """Test 5: Verify Step 32 defined names still exist"""
    print("\nTest 5: Step 32 Defined Names")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_names = list(wb.defined_names.keys())

        missing_names = []
        for expected_name in EXPECTED_STEP32_NAMES:
            if expected_name not in actual_names:
                missing_names.append(expected_name)

        if missing_names:
            print(f"[X] FAIL: Missing Step 32 defined names")
            print(f"   Missing: {missing_names}")
            print(f"   Expected: {EXPECTED_STEP32_NAMES}")
            wb.close()
            return False

        print(f"[OK] PASS: All Step 32 defined names present ({len(EXPECTED_STEP32_NAMES)})")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking Step 32 defined names: {e}")
        return False


def test_home_hyperlinks():
    """Test 6: Verify Home sheet has >= 10 internal hyperlinks"""
    print("\nTest 6: 00_Home Hyperlinks")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        if "00_Home" not in wb.sheetnames:
            print(f"[X] FAIL: 00_Home sheet not found")
            wb.close()
            return False

        ws = wb["00_Home"]

        # Count hyperlinks
        hyperlink_count = 0
        for row in ws.iter_rows(min_row=1, max_row=100):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Check for HYPERLINK formula
                    if "HYPERLINK" in cell.value and "#" in cell.value:
                        hyperlink_count += 1

        if hyperlink_count < 10:
            print(f"[X] FAIL: Not enough hyperlinks in 00_Home")
            print(f"   Found: {hyperlink_count}, Expected: >= 10")
            wb.close()
            return False

        print(f"[OK] PASS: Found {hyperlink_count} hyperlinks in 00_Home (>= 10)")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking hyperlinks: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_total_sheets():
    """Test 7: Verify total sheets >= 17"""
    print("\nTest 7: Total Sheets Count")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        total_sheets = len(wb.sheetnames)

        if total_sheets < 17:
            print(f"[X] FAIL: Not enough sheets")
            print(f"   Found: {total_sheets}, Expected: >= 17")
            wb.close()
            return False

        print(f"[OK] PASS: Total sheets: {total_sheets} (>= 17)")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking sheets count: {e}")
        return False


def test_tables_not_modified():
    """Test 8: Verify tables still exist (from Step 31)"""
    print("\nTest 8: Tables Integrity")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        table_count = 0

        expected_tables = [
            "01Recetas",
            "02RecetaVersiones",
            "03Lotes",
            "04LoteMediciones",
            "05ItemsInventario",
            "06MovimientosInventario",
            "07Proveedores",
            "08LotesInsumo",
            "09ConsumosLote",
            "10Productos",
            "11Clientes",
            "12Ventas",
            "13VentasLineas",
            "14Pagos",
            "15FulfillmentVentaLote"
        ]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, 'tables'):
                for table_name in ws.tables.keys():
                    if table_name in expected_tables:
                        table_count += 1

        if table_count < 15:
            print(f"[X] FAIL: Missing tables")
            print(f"   Found: {table_count}, Expected: 15")
            wb.close()
            return False

        print(f"[OK] PASS: All 15 tables present")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking tables: {e}")
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel MVP Tests - Step 33")
    print("=" * 60)

    # Run all tests
    results = []
    results.append(("File Exists", test_file_exists()))
    results.append(("00_Home First", test_home_sheet_exists_and_first()))
    results.append(("98_Ayuda Exists", test_ayuda_sheet_exists()))
    results.append(("Catalogos Exists", test_catalogos_sheet_exists()))
    results.append(("Step 32 Names", test_step32_defined_names()))
    results.append(("Home Hyperlinks", test_home_hyperlinks()))
    results.append(("Total Sheets", test_total_sheets()))
    results.append(("Tables Integrity", test_tables_not_modified()))

    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] PASS" if result else "[X] FAIL"
        print(f"{status}: {test_name}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n*** All tests passed! ***")
        return 0
    else:
        print(f"\n*** {total - passed} test(s) failed ***")
        return 1


if __name__ == "__main__":
    exit(main())
