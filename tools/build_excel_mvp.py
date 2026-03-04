#!/usr/bin/env python
"""
Quelonio Excel MVP Skeleton Builder
Genera plantilla Excel con hojas, tablas y validaciones según EXCEL_MVP_SPEC.md
"""

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# Configuration
OUTPUT_DIR = Path("data/excel")
OUTPUT_FILE = OUTPUT_DIR / "Quelonio_Excel_MVP_Skeleton.xlsx"

# Sheet names and their headers (in order)
SHEETS_CONFIG = {
    "01_Recetas": [
        "receta_id", "nombre", "estado", "fecha_creacion", "notas"
    ],
    "02_RecetaVersiones": [
        "recipe_version_id", "receta_id", "version_num", "og_target", "fg_target",
        "abv_target", "ibu_target", "srm_target", "eficiencia", "rendimiento_litros",
        "merma_porcentaje", "costo_estimado_batch", "costo_estimado_litro",
        "fecha_version", "notas"
    ],
    "03_Lotes": [
        "batch_id", "recipe_version_id", "fecha_inicio", "fecha_coccion",
        "fecha_fermentacion", "fecha_maduracion", "fecha_envasado",
        "volumen_litros", "og_medido", "fg_medido", "abv_medido",
        "estado", "costo_real_batch", "notas"
    ],
    "04_LoteMediciones": [
        "medicion_id", "batch_id", "tipo_medicion", "valor", "fecha"
    ],
    "05_ItemsInventario": [
        "item_id", "nombre", "tipo", "unidad_medida", "stock_actual",
        "stock_minimo", "costo_unitario", "proveedor_id", "estado", "notas"
    ],
    "06_MovimientosInventario": [
        "movimiento_id", "item_id", "tipo_movimiento", "cantidad",
        "fecha", "referencia", "notas"
    ],
    "07_Proveedores": [
        "proveedor_id", "nombre", "contacto", "email"
    ],
    "08_LotesInsumo": [
        "lote_insumo_id", "item_id", "proveedor_id", "fecha_vencimiento"
    ],
    "09_ConsumosLote": [
        "consumo_id", "batch_id", "item_id", "proveedor_id",
        "lote_insumo_id", "cantidad", "fecha"
    ],
    "10_Productos": [
        "producto_id", "nombre", "receta_id", "volumen_ml",
        "precio_venta", "costo_unitario", "margen_porcentaje", "estado"
    ],
    "11_Clientes": [
        "cliente_id", "nombre", "tipo", "canal", "contacto",
        "email", "telefono", "estado"
    ],
    "12_Ventas": [
        "venta_id", "cliente_id", "fecha", "subtotal", "total",
        "estado_pago", "saldo", "notas"
    ],
    "13_VentasLineas": [
        "linea_id", "venta_id", "producto_id", "cantidad",
        "precio_unitario", "subtotal"
    ],
    "14_Pagos": [
        "pago_id", "venta_id", "fecha", "monto", "metodo"
    ],
    "15_FulfillmentVentaLote": [
        "fulfillment_id", "venta_id", "venta_linea_id", "batch_id",
        "cantidad_litros", "fecha"
    ]
}

# Validation lists for enums
VALIDATION_LISTS = {
    "LIST_RECETAS_ESTADO": ["activa", "archivada", "experimental"],
    "LIST_LOTES_ESTADO": ["planificado", "coccion", "fermentacion", "maduracion", "envasado", "completado", "cancelado"],
    "LIST_ITEMS_TIPO": ["insumo", "producto_terminado", "barril"],
    "LIST_ITEMS_UNIDAD_MEDIDA": ["kg", "litros", "unidades"],
    "LIST_ITEMS_ESTADO": ["activo", "inactivo"],
    "LIST_MOVIMIENTOS_TIPO": ["entrada", "salida", "ajuste"],
    "LIST_CLIENTES_TIPO": ["minorista", "mayorista", "directo"],
    "LIST_CLIENTES_CANAL": ["local", "evento", "online"],
    "LIST_VENTAS_ESTADO_PAGO": ["pendiente", "pagado_parcial", "pagado"],
    "LIST_PAGOS_METODO": ["efectivo", "transferencia", "tarjeta"],
    "LIST_PRODUCTOS_ESTADO": ["activo", "inactivo"]
}

# Map of enum columns to validation lists (sheet_name: column_index -> list_name)
# Column indices are 0-based
ENUM_VALIDATIONS = {
    "01_Recetas": {3: "LIST_RECETAS_ESTADO"},
    "03_Lotes": {11: "LIST_LOTES_ESTADO"},
    "05_ItemsInventario": {2: "LIST_ITEMS_TIPO", 3: "LIST_ITEMS_UNIDAD_MEDIDA", 8: "LIST_ITEMS_ESTADO"},
    "06_MovimientosInventario": {2: "LIST_MOVIMIENTOS_TIPO"},
    "10_Productos": {7: "LIST_PRODUCTOS_ESTADO"},
    "11_Clientes": {2: "LIST_CLIENTES_TIPO", 3: "LIST_CLIENTES_CANAL", 7: "LIST_ITEMS_ESTADO"},
    "12_Ventas": {5: "LIST_VENTAS_ESTADO_PAGO"},
    "14_Pagos": {4: "LIST_PAGOS_METODO"}
}


def create_output_directory():
    """Create output directory if it doesn't exist"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {OUTPUT_DIR.absolute()}")


def create_workbook():
    """Create new workbook"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    return wb


def create_sheet(wb, sheet_name, headers):
    """Create sheet with headers, table, and formatting"""
    ws = wb.create_sheet(title=sheet_name)

    # Write headers in row 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = cell.font.copy(bold=True)

    # Freeze pane on row 1
    ws.freeze_panes = "A2"

    # Create table
    if headers:
        last_col = get_column_letter(len(headers))
        table_range = f"A1:{last_col}1"
        table = Table(displayName=sheet_name.replace("_", ""), ref=table_range)

        table_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True
        )
        table.tableStyleInfo = table_style
        ws.add_table(table)

    # Auto-adjust column widths (heuristic based on header length)
    for col_idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return ws


def create_validation_sheet(wb):
    """Create sheet with validation lists"""
    ws = wb.create_sheet(title="99_Listas")
    row = 1

    for list_name, values in VALIDATION_LISTS.items():
        ws.cell(row=row, column=1, value=list_name)
        ws.cell(row=row, column=1).font = ws.cell(row=row, column=1).font.copy(bold=True)

        for i, value in enumerate(values, start=row + 1):
            ws.cell(row=i, column=1, value=value)

        # Define named range for this list
        list_ref = f"99_Listas!${get_column_letter(1)}${row+1}:${get_column_letter(1)}${row+len(values)}"
        defined_name = DefinedName(name=list_name, attr_text=list_ref)
        wb.defined_names.add(defined_name)

        row += len(values) + 2  # Skip to next section

    return ws


def apply_data_validations(wb):
    """Apply data validation to enum columns"""
    for sheet_name, validations in ENUM_VALIDATIONS.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        headers = SHEETS_CONFIG[sheet_name]

        for col_idx, list_name in validations.items():
            if col_idx >= len(headers):
                continue

            col_letter = get_column_letter(col_idx + 1)
            dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
            dv.error = "Valor no válido. Seleccione de la lista."
            dv.errorTitle = "Error de Validación"
            dv.prompt = "Seleccione un valor de la lista."
            dv.promptTitle = "Validación"

            # Apply to column (rows 2 to 10000)
            dv.add(f"{col_letter}2:{col_letter}10000")
            ws.add_data_validation(dv)


def main():
    """Main function to build Excel MVP skeleton"""
    print("=" * 60)
    print("Quelonio Excel MVP Skeleton Builder")
    print("=" * 60)

    # Create output directory
    create_output_directory()

    # Create workbook
    wb = create_workbook()
    print(f"\nCreating workbook with sheets:")

    # Create all data sheets
    for sheet_name, headers in SHEETS_CONFIG.items():
        create_sheet(wb, sheet_name, headers)
        print(f"  - {sheet_name} ({len(headers)} columns)")

    # Create validation sheet
    create_validation_sheet(wb)
    print(f"  - 99_Listas (validation lists)")

    # Apply data validations
    apply_data_validations(wb)
    print(f"\nApplied data validations to enum columns")

    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n{'=' * 60}")
    print(f"SUCCESS! Excel skeleton created:")
    print(f"  Path: {OUTPUT_FILE.absolute()}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    main()
