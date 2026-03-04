"""Build Excel MVP with Dashboard Filters and Alerts (Step 37)

Adds:
- Period selector (7/30/90 days) with named ranges
- Parameterized KPIs by period
- Alerts panel with traffic light status
- Auxiliary calculations in 21_Calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import shutil

# Constants
EXCEL_PATH = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
BACKUP_DIR = os.path.join("data", "excel", "_backups")

# Alert configuration
ALERTS = [
    {
        "name": "ALERT_StockCritico",
        "label": "Stock Crítico",
        "row": 12,
        "col": 1,
        "count_col": 2,
        "status_col": 3,
        "description": "Items con stock <= stock_minimo"
    },
    {
        "name": "ALERT_LotesDemorados",
        "label": "Lotes Demorados (>30d)",
        "row": 13,
        "col": 1,
        "count_col": 2,
        "status_col": 3,
        "description": "Lotes en curso iniciados hace más de 30 días"
    },
    {
        "name": "ALERT_SaldoPendiente",
        "label": "Saldo Pendiente",
        "row": 14,
        "col": 1,
        "count_col": 2,
        "status_col": 3,
        "description": "Ventas con saldo por cobrar > 0"
    }
]


def create_backup():
    """Create backup of existing Excel file with timestamp."""
    if not os.path.exists(EXCEL_PATH):
        return None
    
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"Quelonio_Excel_MVP_Skeleton_step37_{timestamp}.xlsx")
    shutil.copy2(EXCEL_PATH, backup_path)
    print(f"[OK] Backup created: {backup_path}")
    return backup_path


def get_or_create_catalogos_sheet(wb):
    """Get or create Catalogos sheet."""
    if "Catalogos" in wb.sheetnames:
        return wb["Catalogos"]
    
    ws = wb.create_sheet("Catalogos", len(wb.sheetnames))
    
    # Headers
    headers = ["Catalogo", "Valor", "Descripcion"]
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=1 + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.title = "Catalogos"
    print("[OK] Catalogos sheet created")
    return ws


def add_period_catalog(wb):
    """Add period catalog and named range."""
    ws = get_or_create_catalogos_sheet(wb)
    
    # Check if CAT_PeriodoDias already exists
    start_row = None
    for row in range(2, ws.max_row + 1):
        cat_name = ws.cell(row=row, column=1).value
        if cat_name == "CAT_PeriodoDias":
            start_row = row
            break
    
    # Add catalog if not exists
    if start_row is None:
        start_row = ws.max_row + 1
        
        # Add period values
        periodos = [
            ("CAT_PeriodoDias", 7, "Período corto - 7 días"),
            ("CAT_PeriodoDias", 30, "Período estándar - 30 días"),
            ("CAT_PeriodoDias", 90, "Período largo - 90 días")
        ]
        
        for row_idx, (cat, valor, desc) in enumerate(periodos):
            ws.cell(row=start_row + row_idx, column=1, value=cat)
            ws.cell(row=start_row + row_idx, column=2, value=valor)
            ws.cell(row=start_row + row_idx, column=3, value=desc)
        
        # Create named range for period values
        range_str = f"Catalogos!$B${start_row}:$B${start_row + 2}"
        new_dn = openpyxl.workbook.defined_name.DefinedName(
            name="lst_periodo_dias",
            attr_text=range_str
        )
        wb.defined_names.add(new_dn)
        print(f"[OK] Named range lst_periodo_dias created: {range_str}")
    else:
        # Update named range
        range_str = f"Catalogos!$B${start_row}:$B${start_row + 2}"
        
        # Remove existing named range if any
        for dn_name in list(wb.defined_names):
            if dn_name == "lst_periodo_dias":
                del wb.defined_names[dn_name]
        
        new_dn = openpyxl.workbook.defined_name.DefinedName(
            name="lst_periodo_dias",
            attr_text=range_str
        )
        wb.defined_names.add(new_dn)
        print(f"[OK] Named range lst_periodo_dias updated: {range_str}")
    
    return start_row


def add_period_selector_to_dashboard(wb):
    """Add period selector to 20_Dashboard."""
    ws = wb["20_Dashboard"]
    
    # Add selector label
    label_cell = ws.cell(row=3, column=1, value="Período (días):")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    
    # Add selector cell with default value
    selector_cell = ws.cell(row=3, column=2, value=30)
    selector_cell.font = Font(bold=True, color="4472C4")
    selector_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    selector_cell.alignment = Alignment(horizontal="center")
    selector_cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add data validation
    dv = DataValidation(
        type="list",
        formula1="=lst_periodo_dias",
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Seleccione un período de la lista (7, 30, 90)",
        showDropDown=True
    )
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=3, column=2))
    
    # Create named range for selector
    for dn_name in list(wb.defined_names):
        if dn_name == "SEL_PeriodoDias":
            del wb.defined_names[dn_name]
    
    new_dn = openpyxl.workbook.defined_name.DefinedName(
        name="SEL_PeriodoDias",
        attr_text="20_Dashboard!$B$3"
    )
    wb.defined_names.add(new_dn)
    print("[OK] Period selector added to 20_Dashboard!$B$3")
    print("[OK] Named range SEL_PeriodoDias created")


def add_auxiliary_calculations(wb):
    """Add auxiliary calculations to 21_Calc."""
    if "21_Calc" not in wb.sheetnames:
        ws = wb.create_sheet("21_Calc", len(wb.sheetnames))
        ws.title = "21_Calc"
    else:
        ws = wb["21_Calc"]
    
    # Add calculation headers
    headers = ["Variable", "Formula/Valor", "Descripción"]
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=1 + col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add auxiliary calculations
    calc_data = [
        ["CALC_PeriodoDias", "=SEL_PeriodoDias", "Días del período seleccionado"],
        ["CALC_FechaInicio", "=TODAY()-CALC_PeriodoDias", "Fecha inicio del período"],
        ["CALC_LotesUmbralDias", 30, "Umbral de días para lotes demorados"]
    ]
    
    for row_idx, (var, formula, desc) in enumerate(calc_data):
        ws.cell(row=row_idx + 2, column=1, value=var)
        ws.cell(row=row_idx + 2, column=2, value=formula)
        ws.cell(row=row_idx + 2, column=3, value=desc)
        
        # Create named range for CALC variables
        cell_ref = f"21_Calc!$B${row_idx + 2}"
        for dn_name in list(wb.defined_names):
            if dn_name == var:
                del wb.defined_names[dn_name]
        
        new_dn = openpyxl.workbook.defined_name.DefinedName(
            name=var,
            attr_text=cell_ref
        )
        wb.defined_names.add(new_dn)
    
    ws.freeze_panes = ws.cell(row=3, column=1)
    print("[OK] Auxiliary calculations added to 21_Calc")
    print("[OK] Named ranges for CALC variables created")


def update_kpis_with_period_filter(wb):
    """Update KPIs to use period filter."""
    ws = wb["20_Dashboard"]
    
    # Update KPI_Ventas_Periodo (was KPI_Ventas_30d)
    # Find the KPI_Ventas_30d cell and update its formula
    ventas_kpi_row = 6
    ventas_kpi_col = 3
    
    current_formula = ws.cell(row=ventas_kpi_row, column=ventas_kpi_col).value
    if current_formula and isinstance(current_formula, str):
        new_formula = '=SUMIFS(12_Ventas!F:F,12_Ventas!C:C,">="&CALC_FechaInicio)'
        ws.cell(row=ventas_kpi_row, column=ventas_kpi_col, value=new_formula)
        print("[OK] KPI_Ventas_Periodo updated to use CALC_FechaInicio")
    
    # Update KPI_SaldoPorCobrar to optionally filter by period
    saldo_kpi_row = 8
    saldo_kpi_col = 3
    
    current_formula = ws.cell(row=saldo_kpi_row, column=saldo_kpi_col).value
    if current_formula and isinstance(current_formula, str):
        # Filter saldo by date if the sale is within the period
        new_formula = '=SUMIFS(12_Ventas!H:H,12_Ventas!C:C,">="&CALC_FechaInicio)'
        ws.cell(row=saldo_kpi_row, column=saldo_kpi_col, value=new_formula)
        print("[OK] KPI_SaldoPorCobrar updated to use CALC_FechaInicio")


def add_alerts_panel(wb):
    """Add alerts panel to 20_Dashboard."""
    ws = wb["20_Dashboard"]
    
    # Add alerts section header
    alerts_header = ws.cell(row=11, column=1, value="ALERTAS")
    alerts_header.font = Font(size=12, bold=True, color="FFFFFF")
    alerts_header.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    alerts_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=11, start_column=1, end_row=11, end_column=4)
    
    # Add alert rows
    for alert in ALERTS:
        # Label
        label_cell = ws.cell(row=alert["row"], column=alert["col"], value=alert["label"])
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        
        # Count cell
        count_cell = ws.cell(row=alert["row"], column=alert["count_col"], value="")
        count_cell.font = Font(bold=True, size=12)
        count_cell.alignment = Alignment(horizontal="center")
        count_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        count_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Status cell
        status_cell = ws.cell(row=alert["row"], column=alert["status_col"], value="")
        status_cell.font = Font(bold=True, size=10)
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add formulas for alerts
    
    # ALERT_StockCritico: items with stock_actual <= stock_minimo
    ws.cell(row=12, column=2, value='=COUNTIFS(05_ItemsInventario!D:D,"<="&05_ItemsInventario!E:E)')
    
    # ALERT_LotesDemorados: lots in progress with fecha_inicio > 30 days ago
    estados_en_curso = '("coccion","fermentacion","maduracion","envasado","planificado")'
    ws.cell(row=13, column=2, value='=COUNTIFS(03_Lotes!D:D,"coccion",03_Lotes!E:E,"<"&TODAY()-30)+COUNTIFS(03_Lotes!D:D,"fermentacion",03_Lotes!E:E,"<"&TODAY()-30)+COUNTIFS(03_Lotes!D:D,"maduracion",03_Lotes!E:E,"<"&TODAY()-30)+COUNTIFS(03_Lotes!D:D,"envasado",03_Lotes!E:E,"<"&TODAY()-30)+COUNTIFS(03_Lotes!D:D,"planificado",03_Lotes!E:E,"<"&TODAY()-30)')
    
    # ALERT_SaldoPendiente: sales with saldo_calc > 0
    ws.cell(row=14, column=2, value='=COUNTIF(12_Ventas!H:H,">0")')
    
    # Status formulas: IF count > 0 then "ALERTA" else "OK"
    ws.cell(row=12, column=3, value='=IF(B12>0,"ALERTA","OK")')
    ws.cell(row=13, column=3, value='=IF(B13>0,"ALERTA","OK")')
    ws.cell(row=14, column=3, value='=IF(B14>0,"ALERTA","OK")')
    
    # Format status cells with conditional colors
    for alert_row in [12, 13, 14]:
        status_cell = ws.cell(row=alert_row, column=3)
        if status_cell.value and "ALERTA" in str(status_cell.value):
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="FF0000")
        elif status_cell.value and "OK" in str(status_cell.value):
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
    
    # Add alert descriptions column (optional, for reference)
    ws.cell(row=12, column=4, value="Items con stock <= mínimo")
    ws.cell(row=13, column=4, value="Lotes en curso >30 días")
    ws.cell(row=14, column=4, value="Ventas con saldo pendiente")
    
    for row in [12, 13, 14]:
        ws.cell(row=row, column=4).font = Font(italic=True, size=9, color="666666")
    
    print("[OK] Alerts panel added to 20_Dashboard")


def build():
    """Main build function."""
    print("Building Excel MVP - Step 37: Dashboard Filters and Alerts")
    print("-" * 60)
    
    # Create backup
    create_backup()
    
    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Add period catalog and named range
    add_period_catalog(wb)
    
    # Add period selector to dashboard
    add_period_selector_to_dashboard(wb)
    
    # Add auxiliary calculations
    add_auxiliary_calculations(wb)
    
    # Update KPIs to use period filter
    update_kpis_with_period_filter(wb)
    
    # Add alerts panel
    add_alerts_panel(wb)
    
    # Save workbook
    wb.save(EXCEL_PATH)
    print("-" * 60)
    print(f"[OK] Excel MVP saved to: {EXCEL_PATH}")
    print("[OK] Step 37 build complete!")


if __name__ == "__main__":
    build()
