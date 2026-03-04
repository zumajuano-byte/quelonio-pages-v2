#!/usr/bin/env python
"""
Quelonio Excel MVP Builder - Step 34: Seed Data
Agrega datos demo consistentes con FKs y catálogos
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Configuration
OUTPUT_DIR = Path("data/excel")
OUTPUT_FILE = OUTPUT_DIR / "Quelonio_Excel_MVP_Skeleton.xlsx"
BACKUP_DIR = OUTPUT_DIR / "_backups"


# Seed data
SEED_DATA = {
    "01_Recetas": [
        {
            "receta_id": "REC-001",
            "nombre": "IPA Clásica",
            "estado": "activa",
            "fecha_creacion": "2026-01-01",
            "notas": "Receta base IPA con lúpulos Americanos"
        },
        {
            "receta_id": "REC-002",
            "nombre": "Stout Imperial",
            "estado": "activa",
            "fecha_creacion": "2026-01-05",
            "notas": "Stout oscura con notas de café"
        }
    ],
    "02_RecetaVersiones": [
        {
            "recipe_version_id": "REC-001-V01",
            "receta_id": "REC-001",
            "version_num": 1,
            "og_target": 1.060,
            "fg_target": 1.010,
            "abv_target": 6.5,
            "ibu_target": 50,
            "srm_target": 6,
            "eficiencia": 75.0,
            "rendimiento_litros": 500,
            "merma_porcentaje": 5.0,
            "costo_estimado_batch": 15000,
            "costo_estimado_litro": 30,
            "fecha_version": "2026-01-02",
            "notas": "Versión inicial"
        },
        {
            "recipe_version_id": "REC-001-V02",
            "receta_id": "REC-001",
            "version_num": 2,
            "og_target": 1.065,
            "fg_target": 1.012,
            "abv_target": 7.0,
            "ibu_target": 55,
            "srm_target": 6,
            "eficiencia": 78.0,
            "rendimiento_litros": 500,
            "merma_porcentaje": 5.0,
            "costo_estimado_batch": 15500,
            "costo_estimado_litro": 31,
            "fecha_version": "2026-01-10",
            "notas": "Ajuste de lúpulos"
        },
        {
            "recipe_version_id": "REC-002-V01",
            "receta_id": "REC-002",
            "version_num": 1,
            "og_target": 1.080,
            "fg_target": 1.020,
            "abv_target": 7.8,
            "ibu_target": 70,
            "srm_target": 40,
            "eficiencia": 70.0,
            "rendimiento_litros": 300,
            "merma_porcentaje": 8.0,
            "costo_estimado_batch": 20000,
            "costo_estimado_litro": 67,
            "fecha_version": "2026-01-06",
            "notas": "Versión inicial Stout"
        }
    ],
    "10_Productos": [
        {
            "producto_id": "PROD-001",
            "nombre": "IPA Clásica 500ml",
            "receta_id": "REC-001",
            "volumen_ml": 500,
            "precio_venta": 120,
            "costo_unitario": 50,
            "margen_porcentaje": 58.33,
            "estado": "activo"
        },
        {
            "producto_id": "PROD-002",
            "nombre": "Stout Imperial 500ml",
            "receta_id": "REC-002",
            "volumen_ml": 500,
            "precio_venta": 150,
            "costo_unitario": 65,
            "margen_porcentaje": 56.67,
            "estado": "activo"
        }
    ],
    "07_Proveedores": [
        {
            "proveedor_id": "PROV-001",
            "nombre": "Maltería Nacional S.A.",
            "contacto": "Juan Pérez",
            "email": "ventas@malt.es"
        },
        {
            "proveedor_id": "PROV-002",
            "nombre": "Lúpulos del Sur Ltda.",
            "contacto": "María García",
            "email": "compras@lupulos.cl"
        }
    ],
    "05_ItemsInventario": [
        {
            "item_id": "ITEM-001",
            "nombre": "Malta Pilsen",
            "tipo": "malta",
            "unidad_medida": "kg",
            "stock_actual": 500,
            "stock_minimo": 100,
            "costo_unitario": 2.50,
            "proveedor_id": "PROV-001",
            "estado": "activo",
            "notas": "Malta base"
        },
        {
            "item_id": "ITEM-002",
            "nombre": "Lúpulo Cascade",
            "tipo": "luppulo",
            "unidad_medida": "kg",
            "stock_actual": 50,
            "stock_minimo": 10,
            "costo_unitario": 150.00,
            "proveedor_id": "PROV-002",
            "estado": "activo",
            "notas": "Amargor y aroma"
        },
        {
            "item_id": "ITEM-003",
            "nombre": "Levadura US-05",
            "tipo": "levadura",
            "unidad_medida": "kg",
            "stock_actual": 10,
            "stock_minimo": 2,
            "costo_unitario": 500.00,
            "proveedor_id": "PROV-002",
            "estado": "activo",
            "notas": "Fermentación neutra"
        },
        {
            "item_id": "ITEM-004",
            "nombre": "Agua",
            "tipo": "agua",
            "unidad_medida": "litros",
            "stock_actual": 10000,
            "stock_minimo": 2000,
            "costo_unitario": 0.001,
            "proveedor_id": "",
            "estado": "activo",
            "notas": "Agua potable"
        },
        {
            "item_id": "ITEM-005",
            "nombre": "Adjunto Maíz",
            "tipo": "adjunto",
            "unidad_medida": "kg",
            "stock_actual": 200,
            "stock_minimo": 50,
            "costo_unitario": 1.80,
            "proveedor_id": "PROV-001",
            "estado": "activo",
            "notas": "Adjunto para lighten"
        },
        {
            "item_id": "ITEM-006",
            "nombre": "Clarificante",
            "tipo": "quimico",
            "unidad_medida": "kg",
            "stock_actual": 5,
            "stock_minimo": 1,
            "costo_unitario": 800.00,
            "proveedor_id": "PROV-002",
            "estado": "activo",
            "notas": "Para clarificación"
        },
        {
            "item_id": "ITEM-007",
            "nombre": "Botella 500ml",
            "tipo": "envase",
            "unidad_medida": "unidades",
            "stock_actual": 5000,
            "stock_minimo": 1000,
            "costo_unitario": 2.00,
            "proveedor_id": "PROV-001",
            "estado": "activo",
            "notas": "Envase vidrio"
        },
        {
            "item_id": "ITEM-008",
            "nombre": "Etiqueta IPA",
            "tipo": "etiqueta",
            "unidad_medida": "unidades",
            "stock_actual": 3000,
            "stock_minimo": 500,
            "costo_unitario": 0.50,
            "proveedor_id": "PROV-001",
            "estado": "activo",
            "notas": "Etiqueta para botellas"
        },
        {
            "item_id": "ITEM-009",
            "nombre": "Lúpulo Citra",
            "tipo": "luppulo",
            "unidad_medida": "kg",
            "stock_actual": 30,
            "stock_minimo": 5,
            "costo_unitario": 200.00,
            "proveedor_id": "PROV-002",
            "estado": "activo",
            "notas": "Aroma cítrico"
        },
        {
            "item_id": "ITEM-010",
            "nombre": "Tapón Corona",
            "tipo": "envase",
            "unidad_medida": "unidades",
            "stock_actual": 6000,
            "stock_minimo": 1000,
            "costo_unitario": 0.10,
            "proveedor_id": "PROV-001",
            "estado": "activo",
            "notas": "Cierre botellas"
        }
    ],
    "08_LotesInsumo": [
        {
            "lote_insumo_id": "LI-001",
            "item_id": "ITEM-001",
            "proveedor_id": "PROV-001",
            "fecha_vencimiento": "2027-06-30"
        },
        {
            "lote_insumo_id": "LI-002",
            "item_id": "ITEM-002",
            "proveedor_id": "PROV-002",
            "fecha_vencimiento": "2027-12-31"
        },
        {
            "lote_insumo_id": "LI-003",
            "item_id": "ITEM-003",
            "proveedor_id": "PROV-002",
            "fecha_vencimiento": "2026-06-30"
        },
        {
            "lote_insumo_id": "LI-004",
            "item_id": "ITEM-009",
            "proveedor_id": "PROV-002",
            "fecha_vencimiento": "2027-12-31"
        },
        {
            "lote_insumo_id": "LI-005",
            "item_id": "ITEM-001",
            "proveedor_id": "PROV-001",
            "fecha_vencimiento": "2027-12-31"
        }
    ],
    "03_Lotes": [
        {
            "batch_id": "L008",
            "recipe_version_id": "REC-001-V01",
            "fecha_inicio": "2026-01-15",
            "fecha_coccion": "2026-01-15",
            "fecha_fermentacion": "2026-01-16",
            "fecha_maduracion": "2026-01-20",
            "fecha_envasado": "2026-01-25",
            "volumen_litros": 500,
            "og_medido": 1.061,
            "fg_medido": 1.011,
            "abv_medido": 6.6,
            "estado": "completado",
            "costo_real_batch": 14800,
            "notas": "Primera producción IPA V01"
        },
        {
            "batch_id": "L009",
            "recipe_version_id": "REC-002-V01",
            "fecha_inicio": "2026-01-20",
            "fecha_coccion": "2026-01-20",
            "fecha_fermentacion": "2026-01-21",
            "fecha_maduracion": "2026-01-28",
            "fecha_envasado": "",
            "volumen_litros": 300,
            "og_medido": 1.082,
            "fg_medido": 1.021,
            "abv_medido": 8.0,
            "estado": "maduracion",
            "costo_real_batch": 20500,
            "notas": "Primera producción Stout"
        }
    ],
    "09_ConsumosLote": [
        {"consumo_id": "C-001", "batch_id": "L008", "item_id": "ITEM-001", "proveedor_id": "PROV-001", "lote_insumo_id": "LI-001", "cantidad": 100.0, "fecha": "2026-01-15"},
        {"consumo_id": "C-002", "batch_id": "L008", "item_id": "ITEM-002", "proveedor_id": "PROV-002", "lote_insumo_id": "LI-002", "cantidad": 5.0, "fecha": "2026-01-15"},
        {"consumo_id": "C-003", "batch_id": "L008", "item_id": "ITEM-003", "proveedor_id": "PROV-002", "lote_insumo_id": "LI-003", "cantidad": 2.0, "fecha": "2026-01-16"},
        {"consumo_id": "C-004", "batch_id": "L008", "item_id": "ITEM-004", "proveedor_id": "", "lote_insumo_id": "", "cantidad": 500.0, "fecha": "2026-01-15"},
        {"consumo_id": "C-005", "batch_id": "L008", "item_id": "ITEM-006", "proveedor_id": "PROV-002", "lote_insumo_id": "", "cantidad": 0.5, "fecha": "2026-01-25"},
        {"consumo_id": "C-006", "batch_id": "L009", "item_id": "ITEM-001", "proveedor_id": "PROV-001", "lote_insumo_id": "LI-005", "cantidad": 80.0, "fecha": "2026-01-20"},
        {"consumo_id": "C-007", "batch_id": "L009", "item_id": "ITEM-004", "proveedor_id": "", "lote_insumo_id": "", "cantidad": 300.0, "fecha": "2026-01-20"},
        {"consumo_id": "C-008", "batch_id": "L009", "item_id": "ITEM-003", "proveedor_id": "PROV-002", "lote_insumo_id": "LI-003", "cantidad": 1.5, "fecha": "2026-01-21"},
        {"consumo_id": "C-009", "batch_id": "L009", "item_id": "ITEM-006", "proveedor_id": "PROV-002", "lote_insumo_id": "", "cantidad": 0.3, "fecha": "2026-01-28"},
        {"consumo_id": "C-010", "batch_id": "L009", "item_id": "ITEM-009", "proveedor_id": "PROV-002", "lote_insumo_id": "LI-004", "cantidad": 2.0, "fecha": "2026-01-20"}
    ],
    "11_Clientes": [
        {
            "cliente_id": "CLI-001",
            "nombre": "Bar El Lúpulo",
            "tipo": "bar",
            "canal": "distribucion",
            "contacto": "Carlos Ruiz",
            "email": "info@ellupulo.cl",
            "telefono": "+56 9 1234 5678",
            "estado": "activo"
        },
        {
            "cliente_id": "CLI-002",
            "nombre": "Juan Pérez",
            "tipo": "minorista",
            "canal": "taproom",
            "contacto": "Juan Pérez",
            "email": "juan.perez@email.com",
            "telefono": "+56 9 8765 4321",
            "estado": "activo"
        },
        {
            "cliente_id": "CLI-003",
            "nombre": "Supermercado Central",
            "tipo": "mayorista",
            "canal": "distribucion",
            "contacto": "Ana María López",
            "email": "compras@central.com",
            "telefono": "+56 9 1111 2222",
            "estado": "activo"
        }
    ],
    "12_Ventas": [
        {
            "venta_id": "VEN-001",
            "cliente_id": "CLI-001",
            "fecha": "2026-01-26",
            "subtotal": 12000,
            "total": 12000,
            "estado_pago": "pagado",
            "saldo": 0,
            "notas": "Venta al bar"
        },
        {
            "venta_id": "VEN-002",
            "cliente_id": "CLI-002",
            "fecha": "2026-01-27",
            "subtotal": 600,
            "total": 600,
            "estado_pago": "pagado",
            "saldo": 0,
            "notas": "Venta taproom"
        },
        {
            "venta_id": "VEN-003",
            "cliente_id": "CLI-003",
            "fecha": "2026-01-28",
            "subtotal": 30000,
            "total": 30000,
            "estado_pago": "parcial",
            "saldo": 15000,
            "notas": "Venta mayorista con saldo pendiente"
        }
    ],
    "13_VentasLineas": [
        {
            "linea_id": "VL-001",
            "venta_id": "VEN-001",
            "producto_id": "PROD-001",
            "cantidad": 100,
            "precio_unitario": 120,
            "subtotal": 12000
        },
        {
            "linea_id": "VL-002",
            "venta_id": "VEN-002",
            "producto_id": "PROD-001",
            "cantidad": 5,
            "precio_unitario": 120,
            "subtotal": 600
        },
        {
            "linea_id": "VL-003",
            "venta_id": "VEN-003",
            "producto_id": "PROD-001",
            "cantidad": 150,
            "precio_unitario": 120,
            "subtotal": 18000
        },
        {
            "linea_id": "VL-004",
            "venta_id": "VEN-003",
            "producto_id": "PROD-002",
            "cantidad": 80,
            "precio_unitario": 150,
            "subtotal": 12000
        },
        {
            "linea_id": "VL-005",
            "venta_id": "VEN-001",
            "producto_id": "PROD-002",
            "cantidad": 0,
            "precio_unitario": 150,
            "subtotal": 0
        }
    ],
    "15_FulfillmentVentaLote": [
        {
            "fulfillment_id": "FF-001",
            "venta_id": "VEN-001",
            "venta_linea_id": "VL-001",
            "batch_id": "L008",
            "cantidad_litros": 50.0,
            "fecha": "2026-01-27"
        },
        {
            "fulfillment_id": "FF-002",
            "venta_id": "VEN-002",
            "venta_linea_id": "VL-002",
            "batch_id": "L008",
            "cantidad_litros": 2.5,
            "fecha": "2026-01-27"
        },
        {
            "fulfillment_id": "FF-003",
            "venta_id": "VEN-003",
            "venta_linea_id": "VL-003",
            "batch_id": "L008",
            "cantidad_litros": 75.0,
            "fecha": "2026-01-29"
        },
        {
            "fulfillment_id": "FF-004",
            "venta_id": "VEN-003",
            "venta_linea_id": "VL-004",
            "batch_id": "L009",
            "cantidad_litros": 40.0,
            "fecha": "2026-01-29"
        },
        {
            "fulfillment_id": "FF-005",
            "venta_id": "VEN-001",
            "venta_linea_id": "VL-001",
            "batch_id": "L008",
            "cantidad_litros": 30.0,
            "fecha": "2026-01-27"
        }
    ],
    "14_Pagos": [
        {
            "pago_id": "PAG-001",
            "venta_id": "VEN-001",
            "fecha": "2026-01-26",
            "monto": 12000,
            "metodo": "transferencia"
        },
        {
            "pago_id": "PAG-002",
            "venta_id": "VEN-002",
            "fecha": "2026-01-27",
            "monto": 600,
            "metodo": "efectivo"
        },
        {
            "pago_id": "PAG-003",
            "venta_id": "VEN-003",
            "fecha": "2026-01-28",
            "monto": 15000,
            "metodo": "transferencia"
        }
    ]
}


def create_backup():
    """Create backup of existing file with timestamp"""
    if not OUTPUT_FILE.exists():
        print(f"No existing file to backup: {OUTPUT_FILE}")
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"Quelonio_Excel_MVP_Skeleton_step34_{timestamp}.xlsx"
    backup_path = BACKUP_DIR / backup_name

    shutil.copy2(OUTPUT_FILE, backup_path)
    print(f"Backup created: {backup_path}")
    return backup_path


def load_wb(path):
    """Load workbook with error handling"""
    if not Path(path).exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    print(f"Loading workbook: {path}")
    wb = load_workbook(path, data_only=False)
    return wb


def get_header_map(ws):
    """Get mapping of header names to column indices (0-based)"""
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=0):
        if cell.value:
            header_map[cell.value] = col_idx
    return header_map


def get_existing_ids(ws, id_column):
    """Get set of existing IDs from a column"""
    ids = set()
    header_map = get_header_map(ws)

    if id_column not in header_map:
        return ids

    col_idx = header_map[id_column]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx]:
            ids.add(row[col_idx])

    return ids


def find_last_table_row(ws):
    """Find the last row with data in the table"""
    return ws.max_row if ws.max_row > 1 else 1


def append_rows(ws, data_rows):
    """Append rows to worksheet"""
    if not data_rows:
        return 0

    header_map = get_header_map(ws)
    headers = list(header_map.keys())

    start_row = find_last_table_row(ws) + 1

    for row_data in data_rows:
        for header, value in row_data.items():
            if header in header_map:
                col_idx = header_map[header]
                ws.cell(row=start_row, column=col_idx + 1, value=value)
        start_row += 1

    return len(data_rows)


def seed_sheet(wb, sheet_name, seed_data, id_column=None):
    """Seed a sheet with data, avoiding duplicates"""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet {sheet_name} not found, skipping")
        return 0

    ws = wb[sheet_name]
    added = 0

    # Get existing IDs if id_column is specified
    existing_ids = set()
    if id_column:
        existing_ids = get_existing_ids(ws, id_column)
        print(f"  Existing {id_column}s in {sheet_name}: {len(existing_ids)}")

    # Filter out duplicates
    rows_to_add = []
    for row in seed_data:
        if id_column and row.get(id_column) in existing_ids:
            print(f"  Skipping duplicate {id_column}: {row.get(id_column)}")
            continue
        rows_to_add.append(row)

    if rows_to_add:
        added = append_rows(ws, rows_to_add)
        print(f"  Added {added} rows to {sheet_name}")

        # Update table range if exists
        if hasattr(ws, 'tables') and len(ws.tables) > 0:
            table_name = list(ws.tables.keys())[0]
            table = ws.tables[table_name]

            # Expand table to include new rows
            last_row = find_last_table_row(ws)
            current_ref = table.ref

            import re
            match = re.search(r':([A-Z]+)(\d+)$', current_ref)
            if match:
                end_col = match.group(1)
                new_ref = f"{current_ref.split(':')[0]}:{end_col}{last_row}"
                table.ref = new_ref
    else:
        print(f"  No new rows to add to {sheet_name}")

    return added


def main():
    """Main function to seed Excel MVP Step 34"""
    print("=" * 60)
    print("Quelonio Excel MVP Builder - Step 34: Seed Data")
    print("=" * 60)

    # Create backup
    backup_path = create_backup()

    # Load workbook
    try:
        wb = load_wb(OUTPUT_FILE)
    except Exception as e:
        print(f"\nERROR: Failed to load workbook: {e}")
        return 1

    print(f"  Loaded {len(wb.sheetnames)} sheets\n")

    # Seed all tables
    total_added = 0

    # 01_Recetas (id_column: receta_id)
    total_added += seed_sheet(wb, "01_Recetas", SEED_DATA["01_Recetas"], "receta_id")

    # 02_RecetaVersiones (id_column: recipe_version_id)
    total_added += seed_sheet(wb, "02_RecetaVersiones", SEED_DATA["02_RecetaVersiones"], "recipe_version_id")

    # 10_Productos (id_column: producto_id)
    total_added += seed_sheet(wb, "10_Productos", SEED_DATA["10_Productos"], "producto_id")

    # 07_Proveedores (id_column: proveedor_id)
    total_added += seed_sheet(wb, "07_Proveedores", SEED_DATA["07_Proveedores"], "proveedor_id")

    # 05_ItemsInventario (id_column: item_id)
    total_added += seed_sheet(wb, "05_ItemsInventario", SEED_DATA["05_ItemsInventario"], "item_id")

    # 08_LotesInsumo (id_column: lote_insumo_id)
    total_added += seed_sheet(wb, "08_LotesInsumo", SEED_DATA["08_LotesInsumo"], "lote_insumo_id")

    # 03_Lotes (id_column: batch_id)
    total_added += seed_sheet(wb, "03_Lotes", SEED_DATA["03_Lotes"], "batch_id")

    # 09_ConsumosLote (id_column: consumo_id)
    total_added += seed_sheet(wb, "09_ConsumosLote", SEED_DATA["09_ConsumosLote"], "consumo_id")

    # 11_Clientes (id_column: cliente_id)
    total_added += seed_sheet(wb, "11_Clientes", SEED_DATA["11_Clientes"], "cliente_id")

    # 12_Ventas (id_column: venta_id)
    total_added += seed_sheet(wb, "12_Ventas", SEED_DATA["12_Ventas"], "venta_id")

    # 13_VentasLineas (id_column: linea_id)
    total_added += seed_sheet(wb, "13_VentasLineas", SEED_DATA["13_VentasLineas"], "linea_id")

    # 15_FulfillmentVentaLote (id_column: fulfillment_id)
    total_added += seed_sheet(wb, "15_FulfillmentVentaLote", SEED_DATA["15_FulfillmentVentaLote"], "fulfillment_id")

    # 14_Pagos (id_column: pago_id)
    total_added += seed_sheet(wb, "14_Pagos", SEED_DATA["14_Pagos"], "pago_id")

    try:
        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Excel MVP Step 34 Seed completed:")
        print(f"  Path: {OUTPUT_FILE.absolute()}")
        print(f"  Total rows added: {total_added}")
        print(f"  Backup: {backup_path}")
        print(f"{'=' * 60}\n")
        return 0

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        wb.close()


if __name__ == "__main__":
    exit(main())
