#!/usr/bin/env python
"""
Quelonio Excel MVP Builder - Step 32
Agrega: Catalogos, FKs, Data Validation, Columnas Calculadas
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, column_index_from_string

# Configuration
OUTPUT_DIR = Path("data/excel")
OUTPUT_FILE = OUTPUT_DIR / "Quelonio_Excel_MVP_Skeleton.xlsx"
BACKUP_DIR = OUTPUT_DIR / "_backups"

# Catalog lists
CATALOGS = {
    "CAT_LoteEstado": ["planificado", "coccion", "fermentacion", "maduracion", "envasado", "completado", "cancelado"],
    "CAT_MovTipo": ["entrada", "salida", "ajuste", "consumo", "produccion", "venta"],
    "CAT_PagoEstado": ["pendiente", "pagado", "parcial"],
    "CAT_ClienteTipo": ["minorista", "mayorista", "bar", "distribuidor", "otro"],
    "CAT_Canal": ["taproom", "bar", "distribucion", "online", "evento", "otro"],
    "CAT_ItemTipo": ["malta", "luppulo", "levadura", "agua", "adjunto", "quimico", "envase", "etiqueta", "servicio", "otro"]
}

# Defined names for catalogs (name -> catalog name in Catalogos sheet)
CATALOG_DEFINED_NAMES = {
    "lst_lote_estado": "CAT_LoteEstado",
    "lst_mov_tipo": "CAT_MovTipo",
    "lst_pago_estado": "CAT_PagoEstado",
    "lst_cliente_tipo": "CAT_ClienteTipo",
    "lst_canal": "CAT_Canal",
    "lst_item_tipo": "CAT_ItemTipo"
}

# Defined names for ID columns (FK targets)
ID_DEFINED_NAMES = {
    "lst_receta_id": ("01_Recetas", "01Recetas", "receta_id"),
    "lst_recipe_version_id": ("02_RecetaVersiones", "02RecetaVersiones", "recipe_version_id"),
    "lst_batch_id": ("03_Lotes", "03Lotes", "batch_id"),
    "lst_item_id": ("05_ItemsInventario", "05ItemsInventario", "item_id"),
    "lst_lote_insumo_id": ("08_LotesInsumo", "08LotesInsumo", "lote_insumo_id"),
    "lst_venta_id": ("12_Ventas", "12Ventas", "venta_id"),
    "lst_linea_id": ("13_VentasLineas", "13VentasLineas", "linea_id"),
    "lst_producto_id": ("10_Productos", "10Productos", "producto_id")
}

# Enum validations (sheet_name: column_name -> list_name)
ENUM_VALIDATIONS = {
    "03_Lotes": {"estado": "lst_lote_estado"},
    "06_MovimientosInventario": {"tipo_movimiento": "lst_mov_tipo"},
    "05_ItemsInventario": {"tipo": "lst_item_tipo"},
    "11_Clientes": {"tipo": "lst_cliente_tipo", "canal": "lst_canal"},
    "12_Ventas": {"estado_pago": "lst_pago_estado"}
}

# FK validations (sheet_name: column_name -> id_list_name)
FK_VALIDATIONS = {
    "02_RecetaVersiones": {"receta_id": "lst_receta_id"},
    "03_Lotes": {"recipe_version_id": "lst_recipe_version_id"},
    "13_VentasLineas": {"venta_id": "lst_venta_id", "producto_id": "lst_producto_id"},
    "14_Pagos": {"venta_id": "lst_venta_id"},
    "09_ConsumosLote": {"batch_id": "lst_batch_id", "item_id": "lst_item_id", "lote_insumo_id": "lst_lote_insumo_id"},
    "15_FulfillmentVentaLote": {"venta_linea_id": "lst_linea_id", "batch_id": "lst_batch_id"}
}

# Calculated columns to add (sheet_name: column_name -> formula)
CALCULATED_COLUMNS = {
    "03_Lotes": {
        "abv_estimado": '=IF(AND([@og_medido]<>"",[@fg_medido]<>""), ([@og_medido]-[@fg_medido])*131.25, "")'
    },
    "12_Ventas": {
        "total_calc": '=IF([@venta_id]="","",SUMIF(VentasLineas[venta_id],[@venta_id],VentasLineas[subtotal]))',
        "pagos_calc": '=IF([@venta_id]="","",SUMIF(Pagos[venta_id],[@venta_id],Pagos[monto]))',
        "saldo_calc": '=IF([@venta_id]="","",[@total_calc]-[@pagos_calc])'
    }
}


def create_backup():
    """Create backup of existing file with timestamp"""
    if not OUTPUT_FILE.exists():
        print(f"No existing file to backup: {OUTPUT_FILE}")
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"Quelonio_Excel_MVP_Skeleton_step32_{timestamp}.xlsx"
    backup_path = BACKUP_DIR / backup_name

    shutil.copy2(OUTPUT_FILE, backup_path)
    print(f"Backup created: {backup_path}")
    return backup_path


def load_workbook_safe():
    """Load workbook with error handling"""
    if not OUTPUT_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {OUTPUT_FILE}")

    print(f"Loading workbook: {OUTPUT_FILE}")
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    return wb


def create_catalogos_sheet(wb):
    """Create or update Catalogos sheet with catalog lists"""
    print("\nCreating/updating Catalogos sheet...")

    sheet_name = "Catalogos"

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        print(f"  Removed existing {sheet_name} sheet")

    ws = wb.create_sheet(title=sheet_name, index=len(wb.sheetnames))

    row = 1
    catalog_ranges = {}

    for catalog_name, values in CATALOGS.items():
        ws.cell(row=row, column=1, value=catalog_name)
        ws.cell(row=row, column=1).font = ws.cell(row=row, column=1).font.copy(bold=True)

        for i, value in enumerate(values, start=row + 1):
            ws.cell(row=i, column=1, value=value)

        catalog_ranges[catalog_name] = f"${get_column_letter(1)}${row+1}:${get_column_letter(1)}${row+len(values)}"
        row += len(values) + 2

    print(f"  Created {sheet_name} with {len(CATALOGS)} catalogs")
    return ws, catalog_ranges


def create_catalog_defined_names(wb, catalog_ranges):
    """Create defined names for catalog lists"""
    print("\nCreating defined names for catalogs...")

    for name, catalog_name in CATALOG_DEFINED_NAMES.items():
        if catalog_name not in catalog_ranges:
            print(f"  WARNING: Catalog {catalog_name} not found in ranges")
            continue

        range_ref = f"Catalogos!{catalog_ranges[catalog_name]}"

        # Remove if exists
        if name in wb.defined_names:
            del wb.defined_names[name]

        defined_name = DefinedName(name=name, attr_text=range_ref)
        wb.defined_names.add(defined_name)
        print(f"  Created: {name} -> {range_ref}")

    print(f"  Created {len(CATALOG_DEFINED_NAMES)} catalog defined names")


def create_id_defined_names(wb):
    """Create defined names for ID columns (FK targets)"""
    print("\nCreating defined names for ID columns (FKs)...")

    for name, (sheet_name, table_name, column_name) in ID_DEFINED_NAMES.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {sheet_name} not found, skipping {name}")
            continue

        ws = wb[sheet_name]

        if not hasattr(ws, 'tables') or table_name not in ws.tables:
            print(f"  WARNING: Table {table_name} not found in {sheet_name}, skipping {name}")
            continue

        table = ws.tables[table_name]
        table_ref = table.ref

        # Find column index
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        if column_name not in headers:
            print(f"  WARNING: Column {column_name} not found in {sheet_name}, skipping {name}")
            continue

        col_idx = headers.index(column_name) + 1
        col_letter = get_column_letter(col_idx)

        # Parse table ref to get range
        parts = table_ref.split(':')
        start_cell = parts[0]
        start_row = int(''.join(filter(str.isdigit, start_cell)))

        # Create reference excluding header: column[2]:column[last]
        end_cell = parts[1] if len(parts) > 1 else start_cell
        end_row = int(''.join(filter(str.isdigit, end_cell)))

        range_ref = f"{sheet_name}!{col_letter}{start_row+1}:{col_letter}{end_row}"

        # Remove if exists
        if name in wb.defined_names:
            del wb.defined_names[name]

        defined_name = DefinedName(name=name, attr_text=range_ref)
        wb.defined_names.add(defined_name)
        print(f"  Created: {name} -> {range_ref}")

    print(f"  Created {len(ID_DEFINED_NAMES)} ID defined names")


def get_column_index_by_name(ws, column_name):
    """Get 0-based column index by header name"""
    for col_idx, cell in enumerate(ws[1], start=0):
        if cell.value == column_name:
            return col_idx
    return None


def apply_data_validations(wb):
    """Apply data validation (enums and FKs)"""
    print("\nApplying data validations...")

    validation_count = 0

    # Apply enum validations
    for sheet_name, validations in ENUM_VALIDATIONS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {sheet_name} not found, skipping enum validations")
            continue

        ws = wb[sheet_name]

        for column_name, list_name in validations.items():
            col_idx = get_column_index_by_name(ws, column_name)
            if col_idx is None:
                print(f"  WARNING: Column {column_name} not found in {sheet_name}, skipping validation")
                continue

            col_letter = get_column_letter(col_idx + 1)
            dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
            dv.error = "Valor no válido. Seleccione de la lista."
            dv.errorTitle = "Error de Validación"
            dv.prompt = "Seleccione un valor de la lista."
            dv.promptTitle = "Validación"

            dv.add(f"{col_letter}2:{col_letter}5000")
            ws.add_data_validation(dv)
            validation_count += 1
            print(f"  Added enum validation: {sheet_name}[{column_name}] -> {list_name}")

    # Apply FK validations
    for sheet_name, validations in FK_VALIDATIONS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {sheet_name} not found, skipping FK validations")
            continue

        ws = wb[sheet_name]

        for column_name, list_name in validations.items():
            col_idx = get_column_index_by_name(ws, column_name)
            if col_idx is None:
                print(f"  WARNING: Column {column_name} not found in {sheet_name}, skipping validation")
                continue

            col_letter = get_column_letter(col_idx + 1)
            dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
            dv.error = "Valor no válido. Seleccione de la lista."
            dv.errorTitle = "Error de Validación"
            dv.prompt = "Seleccione un valor de la lista."
            dv.promptTitle = "Validación"

            dv.add(f"{col_letter}2:{col_letter}5000")
            ws.add_data_validation(dv)
            validation_count += 1
            print(f"  Added FK validation: {sheet_name}[{column_name}] -> {list_name}")

    print(f"  Applied {validation_count} data validations")


def add_calculated_columns(wb):
    """Add calculated columns to tables if they don't exist"""
    print("\nAdding calculated columns...")

    for sheet_name, columns in CALCULATED_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {sheet_name} not found, skipping calculated columns")
            continue

        ws = wb[sheet_name]

        if not hasattr(ws, 'tables') or len(ws.tables) == 0:
            print(f"  WARNING: No tables found in {sheet_name}, skipping calculated columns")
            continue

        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]

        # First, collect all columns to add
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        columns_to_add = []

        for column_name, formula in columns.items():
            if column_name not in headers:
                columns_to_add.append((column_name, formula))

        if not columns_to_add:
            print(f"  All calculated columns already exist in {sheet_name}")
            continue

        # Add all new columns
        start_col_idx = len(headers)

        for offset, (column_name, formula) in enumerate(columns_to_add):
            col_idx = start_col_idx + offset + 1
            col_letter = get_column_letter(col_idx)

            # Add header
            ws.cell(row=1, column=col_idx, value=column_name)
            header_cell = ws.cell(row=1, column=col_idx)
            from openpyxl.styles import Font
            header_cell.font = Font(bold=True)

            # Add formula to all data rows (2 to 5000)
            for row_idx in range(2, 5001):
                ws.cell(row=row_idx, column=col_idx, value=formula)

            print(f"  Added calculated column {column_name} to {sheet_name}")

        # Update table reference to include all new columns
        final_col_idx = len(headers) + len(columns_to_add)
        final_col_letter = get_column_letter(final_col_idx)

        current_ref = table.ref
        parts = current_ref.split(':')
        start_cell = parts[0]
        end_cell = parts[1] if len(parts) > 1 else start_cell

        # Parse end cell to get row number
        import re
        end_row_match = re.search(r'\d+', end_cell)
        if end_row_match:
            end_row = end_row_match.group()
            new_ref = f"{start_cell}:{final_col_letter}{end_row}"
            table.ref = new_ref


def main():
    """Main function to build Excel MVP Step 32"""
    print("=" * 60)
    print("Quelonio Excel MVP Builder - Step 32")
    print("=" * 60)

    # Create backup
    backup_path = create_backup()

    # Load existing workbook
    try:
        wb = load_workbook_safe()
    except Exception as e:
        print(f"\nERROR: Failed to load workbook: {e}")
        return 1

    print(f"  Loaded {len(wb.sheetnames)} sheets")

    try:
        # Create/update Catalogos sheet
        catalogos_ws, catalog_ranges = create_catalogos_sheet(wb)

        # Create catalog defined names
        create_catalog_defined_names(wb, catalog_ranges)

        # Create ID defined names (FKs)
        create_id_defined_names(wb)

        # Apply data validations
        apply_data_validations(wb)

        # Add calculated columns
        add_calculated_columns(wb)

        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Excel MVP Step 32 completed:")
        print(f"  Path: {OUTPUT_FILE.absolute()}")
        print(f"  Sheets: {len(wb.sheetnames)}")
        print(f"  Backup: {backup_path}")
        print(f"{'=' * 60}\n")
        return 0

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        wb.close()


if __name__ == "__main__":
    exit(main())
