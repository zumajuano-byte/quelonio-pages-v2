#!/usr/bin/env python
"""
Quelonio Excel MVP Builder - Step 35: Hygiene
Corrige rangos de tablas infladas y ajusta conteo real de filas
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

# Configuration
OUTPUT_DIR = Path("data/excel")
OUTPUT_FILE = OUTPUT_DIR / "Quelonio_Excel_MVP_Skeleton.xlsx"
BACKUP_DIR = OUTPUT_DIR / "_backups"

# ID column for each sheet
SHEET_ID_COLUMNS = {
    "01_Recetas": "receta_id",
    "02_RecetaVersiones": "recipe_version_id",
    "03_Lotes": "batch_id",
    "04_LoteMediciones": "medicion_id",
    "05_ItemsInventario": "item_id",
    "06_MovimientosInventario": "movimiento_id",
    "07_Proveedores": "proveedor_id",
    "08_LotesInsumo": "lote_insumo_id",
    "09_ConsumosLote": "consumo_id",
    "10_Productos": "producto_id",
    "11_Clientes": "cliente_id",
    "12_Ventas": "venta_id",
    "13_VentasLineas": "linea_id",
    "14_Pagos": "pago_id",
    "15_FulfillmentVentaLote": "fulfillment_id"
}

# Sheets to process (business sheets only)
BUSINESS_SHEETS = list(SHEET_ID_COLUMNS.keys())


def create_backup():
    """Create backup of existing file with timestamp"""
    if not OUTPUT_FILE.exists():
        print(f"No existing file to backup: {OUTPUT_FILE}")
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"Quelonio_Excel_MVP_Skeleton_step35_{timestamp}.xlsx"
    backup_path = BACKUP_DIR / backup_name

    shutil.copy2(OUTPUT_FILE, backup_path)
    print(f"Backup created: {backup_path}")
    return backup_path


def load_wb(path):
    """Load workbook with error handling"""
    if not Path(path).exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    print(f"Loading workbook: {path}")
    wb = load_workbook(path, data_only=False)
    return wb


def get_header_map(ws):
    """Get mapping of header names to column indices (0-based)"""
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=0):
        if cell.value:
            header_map[cell.value] = col_idx
    return header_map


def get_last_data_row(ws, id_column):
    """Find last row where ID column is not empty"""
    header_map = get_header_map(ws)

    if id_column not in header_map:
        print(f"  WARNING: ID column '{id_column}' not found, using max_row - 1")
        return max(1, ws.max_row - 1)

    col_idx = header_map[id_column]

    # Search from bottom up for first non-empty ID
    for row in range(ws.max_row, 1, -1):
        cell_value = ws.cell(row=row, column=col_idx + 1).value
        if cell_value is not None and cell_value != "":
            return row

    # If no data found, return 1 (header row only)
    return 1


def shrink_table_range(ws, last_data_row):
    """Shrink table reference to actual data range"""
    if not hasattr(ws, 'tables') or len(ws.tables) == 0:
        return False, "No table found"

    table_name = list(ws.tables.keys())[0]
    table = ws.tables[table_name]

    current_ref = table.ref

    # Parse current reference to get start cell and columns
    # Format: "A1:Z5000" or similar
    parts = current_ref.split(':')
    start_cell = parts[0]

    # Get the column letter from the end cell
    if len(parts) > 1:
        end_cell_full = parts[1]
        # Extract column letter (letters before numbers)
        end_col_match = re.match(r'^([A-Z]+)', end_cell_full)
        if end_col_match:
            end_col_letter = end_col_match.group(1)
            # Build new reference with correct row
            new_ref = f"{start_cell}:{end_col_letter}{last_data_row}"
            table.ref = new_ref
            return True, f"Updated from {current_ref} to {new_ref}"

    return False, "Could not parse table reference"


def move_data_to_correct_rows(ws, id_column, data_rows, calculated_formulas):
    """Move data rows to correct position starting at row 2, preserving all cell values/formulas"""
    if not data_rows:
        return False

    # Delete all data rows (2 to max_row)
    ws.delete_rows(2, ws.max_row - 1)

    # Write data back starting at row 2
    for row_idx, data_row in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(data_row, start=1):
            # Check if this is a calculated column - use formula instead of value
            if col_idx - 1 in calculated_formulas:
                ws.cell(row=row_idx, column=col_idx, value=calculated_formulas[col_idx - 1])
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)

    return True


def process_sheet_hygiene(wb, sheet_name):
    """Process hygiene for a single sheet - actually shrink by moving data"""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet {sheet_name} not found, skipping")
        return False, 0, 0

    ws = wb[sheet_name]
    id_column = SHEET_ID_COLUMNS.get(sheet_name)

    # Get actual max row (before hygiene)
    actual_max_row = ws.max_row

    # Collect all data rows (rows with non-empty IDs)
    header_map = get_header_map(ws)

    if not id_column or id_column not in header_map:
        print(f"  WARNING: No ID column found for {sheet_name}")
        return False, actual_max_row, 1

    col_idx = header_map[id_column]

    # Collect calculated column formulas from first few rows (up to row 100)
    calculated_formulas = {}
    for header_name, col_idx in header_map.items():
        if 'estimado' in header_name.lower() or '_calc' in header_name.lower():
            # Get formula from first few rows that has it
            for row in range(2, min(100, actual_max_row + 1)):
                cell = ws.cell(row=row, column=col_idx + 1)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    calculated_formulas[col_idx] = cell.value
                    break

    # Collect ALL rows from 2 to max_row (not just ones with non-empty IDs)
    # This preserves formulas from empty validation rows where they might exist
    data_rows = []
    for row in range(2, actual_max_row + 1):
        # Get entire row data (preserve formulas in calculated columns)
        row_data = []
        for c in range(1, len(header_map) + 1):
            row_data.append(ws.cell(row=row, column=c).value)
        data_rows.append(row_data)

    last_data_row_original = 1 + len(data_rows)

    # Move data to correct position (starting at row 2) - only moves data rows
    moved = move_data_to_correct_rows(ws, id_column, data_rows, calculated_formulas)

    # Find new last data row (should be 1 + len(data_rows))
    new_last_data_row = 1 + len(data_rows)

    # Get last column letter
    last_col_idx = len(header_map)
    last_col_letter = get_column_letter(last_col_idx)

    # Shrink table reference
    table_shrunk = False
    table_message = ""
    if hasattr(ws, 'tables') and len(ws.tables) > 0 and len(data_rows) > 0:
        table_name = list(ws.tables.keys())[0]
        table = ws.tables[table_name]

        # Build new reference
        new_ref = f"A1:{last_col_letter}{new_last_data_row}"
        old_ref = table.ref

        table.ref = new_ref
        table_shrunk = True
        table_message = f"Shrunk from {old_ref} to {new_ref} (moved {len(data_rows)} rows)"

    status_msg = f"{sheet_name}: "
    status_msg += f"max_row={actual_max_row} -> {ws.max_row}, "
    status_msg += f"data_rows={len(data_rows)}, "
    status_msg += f"new_last_row={new_last_data_row}"
    if table_shrunk:
        status_msg += f" | {table_message}"
    elif len(data_rows) == 0:
        status_msg += " | No data to move"

    print(f"  {status_msg}")

    return table_shrunk, actual_max_row, new_last_data_row


def main():
    """Main function to clean Excel MVP Step 35"""
    print("=" * 60)
    print("Quelonio Excel MVP Builder - Step 35: Hygiene")
    print("=" * 60)

    # Create backup
    backup_path = create_backup()

    # Load workbook
    try:
        wb = load_wb(OUTPUT_FILE)
    except Exception as e:
        print(f"\nERROR: Failed to load workbook: {e}")
        return 1

    print(f"  Loaded {len(wb.sheetnames)} sheets\n")

    # Process all business sheets
    print("Processing sheet hygiene...")
    print("-" * 60)

    tables_shrunk = 0
    total_before_rows = 0
    total_after_rows = 0

    for sheet_name in BUSINESS_SHEETS:
        shrunk, before, after = process_sheet_hygiene(wb, sheet_name)
        if shrunk:
            tables_shrunk += 1
        total_before_rows += before
        total_after_rows += after

    print("-" * 60)
    print(f"Summary:")
    print(f"  Tables shrunk: {tables_shrunk}")
    print(f"  Total rows before: {total_before_rows}")
    print(f"  Total rows after: {total_after_rows}")
    print(f"  Rows eliminated: {total_before_rows - total_after_rows}")

    try:
        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Excel MVP Step 35 Hygiene completed:")
        print(f"  Path: {OUTPUT_FILE.absolute()}")
        print(f"  Backup: {backup_path}")
        print(f"{'=' * 60}\n")
        return 0

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        wb.close()


if __name__ == "__main__":
    exit(main())
