#!/usr/bin/env python
"""
Test script for Excel MVP Step 34
Verifica: Seed data, Integridad FKs, Export CSV
"""

from pathlib import Path
from openpyxl import load_workbook

# Configuration
OUTPUT_FILE = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")
EXPORT_DIR = Path("data/excel/exports")

# Minimum row counts expected
MIN_ROW_COUNTS = {
    "01_Recetas": 2,
    "02_RecetaVersiones": 3,
    "03_Lotes": 2,
    "05_ItemsInventario": 10,
    "07_Proveedores": 2,
    "08_LotesInsumo": 5,
    "09_ConsumosLote": 10,
    "10_Productos": 2,
    "11_Clientes": 3,
    "12_Ventas": 3,
    "13_VentasLineas": 5,
    "14_Pagos": 3,
    "15_FulfillmentVentaLote": 5
}

# FK relationships to validate
FK_RELATIONSHIPS = [
    ("02_RecetaVersiones", "receta_id", "01_Recetas", "receta_id"),
    ("03_Lotes", "recipe_version_id", "02_RecetaVersiones", "recipe_version_id"),
    ("10_Productos", "receta_id", "01_Recetas", "receta_id"),
    ("13_VentasLineas", "venta_id", "12_Ventas", "venta_id"),
    ("13_VentasLineas", "producto_id", "10_Productos", "producto_id"),
    ("14_Pagos", "venta_id", "12_Ventas", "venta_id"),
    ("15_FulfillmentVentaLote", "venta_linea_id", "13_VentasLineas", "linea_id"),
    ("15_FulfillmentVentaLote", "batch_id", "03_Lotes", "batch_id"),
    ("09_ConsumosLote", "batch_id", "03_Lotes", "batch_id"),
    ("09_ConsumosLote", "item_id", "05_ItemsInventario", "item_id"),
    ("09_ConsumosLote", "lote_insumo_id", "08_LotesInsumo", "lote_insumo_id"),
    ("05_ItemsInventario", "proveedor_id", "07_Proveedores", "proveedor_id"),
    ("08_LotesInsumo", "item_id", "05_ItemsInventario", "item_id"),
    ("08_LotesInsumo", "proveedor_id", "07_Proveedores", "proveedor_id"),
    ("12_Ventas", "cliente_id", "11_Clientes", "cliente_id")
]


def test_file_exists():
    """Test 1: Verify Excel file exists"""
    print("\nTest 1: File Existence")
    print("-" * 50)

    if not OUTPUT_FILE.exists():
        print(f"[X] FAIL: File does not exist: {OUTPUT_FILE}")
        return False

    print(f"[OK] PASS: File exists")
    return True


def test_min_row_counts():
    """Test 2: Verify minimum row counts in each table"""
    print("\nTest 2: Minimum Row Counts")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=True)
        all_pass = True

        for sheet_name, min_count in MIN_ROW_COUNTS.items():
            if sheet_name not in wb.sheetnames:
                print(f"[X] FAIL: Sheet '{sheet_name}' not found")
                all_pass = False
                continue

            ws = wb[sheet_name]
            row_count = ws.max_row - 1  # Exclude header

            if row_count < min_count:
                print(f"[X] FAIL: {sheet_name} has {row_count} rows, expected >= {min_count}")
                all_pass = False
            else:
                print(f"[OK] PASS: {sheet_name} has {row_count} rows (>= {min_count})")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking row counts: {e}")
        import traceback
        traceback.print_exc()
        return False


def get_sheet_data(wb, sheet_name, id_column):
    """Get set of IDs from a sheet"""
    if sheet_name not in wb.sheetnames:
        return set()

    ws = wb[sheet_name]

    # Get header map
    header_map = {}
    for col_idx, cell in enumerate(ws[1], start=0):
        if cell.value:
            header_map[cell.value] = col_idx

    if id_column not in header_map:
        return set()

    col_idx = header_map[id_column]
    ids = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_idx]:
            ids.add(row[col_idx])

    return ids


def test_fk_integrity():
    """Test 3: Verify FK integrity across tables"""
    print("\nTest 3: FK Integrity")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=True)
        all_pass = True

        for src_sheet, fk_col, target_sheet, target_col in FK_RELATIONSHIPS:
            # Get all IDs in source column
            src_ids = set()
            ws_src = wb[src_sheet]

            header_map = {}
            for col_idx, cell in enumerate(ws_src[1], start=0):
                if cell.value:
                    header_map[cell.value] = col_idx

            if fk_col not in header_map:
                print(f"[X] FAIL: Column '{fk_col}' not found in {src_sheet}")
                all_pass = False
                continue

            col_idx = header_map[fk_col]
            for row in ws_src.iter_rows(min_row=2, values_only=True):
                val = row[col_idx]
                if val:
                    src_ids.add(val)

            # Get target IDs (allow empty FKs)
            target_ids = get_sheet_data(wb, target_sheet, target_col)

            # Check FK integrity (only non-empty values)
            invalid_fks = []
            for fk_id in src_ids:
                if fk_id and fk_id not in target_ids:
                    invalid_fks.append(fk_id)

            if invalid_fks:
                print(f"[X] FAIL: {src_sheet}[{fk_col}] has {len(invalid_fks)} invalid FKs to {target_sheet}[{target_col}]")
                print(f"  Invalid IDs: {invalid_fks[:5]}...")  # Show first 5
                all_pass = False
            else:
                print(f"[OK] PASS: {src_sheet}[{fk_col}] -> {target_sheet}[{target_col}] ({len(src_ids)} references)")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking FK integrity: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_csv_export_exists():
    """Test 4: Verify CSV export pack exists and has >= 15 files"""
    print("\nTest 4: CSV Export Pack")
    print("-" * 50)

    try:
        # Find latest export folder
        if not EXPORT_DIR.exists():
            print(f"[X] FAIL: Export directory not found: {EXPORT_DIR}")
            return False

        # List all STEP34_pack_* folders
        export_folders = sorted([d for d in EXPORT_DIR.iterdir() if d.is_dir() and d.name.startswith("STEP34_pack_")])

        if not export_folders:
            print(f"[X] FAIL: No STEP34_pack_ folders found in {EXPORT_DIR}")
            return False

        latest_folder = export_folders[-1]
        csv_files = list(latest_folder.glob("*.csv"))

        if len(csv_files) < 15:
            print(f"[X] FAIL: Not enough CSV files found")
            print(f"   Found: {len(csv_files)}, Expected: >= 15")
            print(f"   Folder: {latest_folder}")
            return False

        print(f"[OK] PASS: CSV export pack found")
        print(f"   Folder: {latest_folder}")
        print(f"   CSV files: {len(csv_files)}")
        print(f"   Files: {', '.join([f.name for f in sorted(csv_files)])}")

        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking CSV export: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_steps_31_32_33_compatibility():
    """Test 5: Verify Steps 31/32/33 compatibility"""
    print("\nTest 5: Steps 31/32/33 Compatibility")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        # Check for Step 31 sheets
        step31_sheets = [
            "01_Recetas",
            "02_RecetaVersiones",
            "03_Lotes",
            "04_LoteMediciones",
            "05_ItemsInventario",
            "06_MovimientosInventario",
            "07_Proveedores",
            "08_LotesInsumo",
            "09_ConsumosLote",
            "10_Productos",
            "11_Clientes",
            "12_Ventas",
            "13_VentasLineas",
            "14_Pagos",
            "15_FulfillmentVentaLote",
            "99_Listas"
        ]

        missing_sheets = []
        for sheet in step31_sheets:
            if sheet not in wb.sheetnames:
                missing_sheets.append(sheet)

        if missing_sheets:
            print(f"[X] FAIL: Missing Step 31 sheets")
            print(f"   Missing: {missing_sheets}")
            wb.close()
            return False

        # Check for Step 32 catalog sheet
        if "Catalogos" not in wb.sheetnames:
            print(f"[X] FAIL: Catalogos sheet not found (Step 32)")
            wb.close()
            return False

        # Check for Step 32 defined names (lst_*)
        lst_names = [name for name in wb.defined_names.keys() if name.startswith("lst_")]
        if len(lst_names) < 14:
            print(f"[X] FAIL: Missing Step 32 defined names (lst_*)")
            print(f"   Found: {len(lst_names)}, Expected: 14")
            wb.close()
            return False

        # Check for Step 33 sheets
        step33_sheets = ["00_Home", "98_Ayuda"]
        missing_33 = []
        for sheet in step33_sheets:
            if sheet not in wb.sheetnames:
                missing_33.append(sheet)

        if missing_33:
            print(f"[X] FAIL: Missing Step 33 sheets")
            print(f"   Missing: {missing_33}")
            wb.close()
            return False

        print(f"[OK] PASS: Steps 31/32/33 compatible")
        print(f"   Step 31 sheets: {len(step31_sheets)}")
        print(f"   Step 32 Catalogos: OK")
        print(f"   Step 32 defined names (lst_*): {len(lst_names)}")
        print(f"   Step 33 sheets: {len(step33_sheets)}")

        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking compatibility: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel MVP Tests - Step 34")
    print("=" * 60)

    # Run all tests
    results = []
    results.append(("File Exists", test_file_exists()))
    results.append(("Min Row Counts", test_min_row_counts()))
    results.append(("FK Integrity", test_fk_integrity()))
    results.append(("CSV Export", test_csv_export_exists()))
    results.append(("Steps 31-33 Compatibility", test_steps_31_32_33_compatibility()))

    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] PASS" if result else "[X] FAIL"
        print(f"{status}: {test_name}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n*** All tests passed! ***")
        return 0
    else:
        print(f"\n*** {total - passed} test(s) failed ***")
        return 1


if __name__ == "__main__":
    exit(main())
