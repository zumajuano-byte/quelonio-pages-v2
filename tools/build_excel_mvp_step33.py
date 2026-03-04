#!/usr/bin/env python
"""
Quelonio Excel MVP Builder - Step 33
Agrega UX mínima: hoja 00_Home y 98_Ayuda
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
OUTPUT_DIR = Path("data/excel")
OUTPUT_FILE = OUTPUT_DIR / "Quelonio_Excel_MVP_Skeleton.xlsx"
BACKUP_DIR = OUTPUT_DIR / "_backups"

# Sheet names for navigation (to link from Home)
NAVIGATION_SHEETS = [
    "01_Recetas",
    "02_RecetaVersiones",
    "03_Lotes",
    "05_ItemsInventario",
    "08_LotesInsumo",
    "09_ConsumosLote",
    "10_Productos",
    "11_Clientes",
    "12_Ventas",
    "13_VentasLineas",
    "14_Pagos",
    "15_FulfillmentVentaLote"
]

# Help content
AYUDA_CONTENT = [
    ("CONVENCION DE IDs", [
        "Los IDs son texto y se auto-generan.",
        "Formato sugerido: PREFIX_XXXX (ej: RECETA_0001)",
        "Usar ID cuando se necesite referencia única.",
        "No modificar IDs existentes para mantener integridad."
    ]),
    ("ORDEN RECOMENDADO DE CARGA", [
        "1. ItemsInventario (materias primas)",
        "2. Proveedores",
        "3. LotesInsumo (lotes de materias primas)",
        "4. Recetas",
        "5. RecetaVersiones",
        "6. Clientes",
        "7. Productos",
        "8. Lotes (producción)",
        "9. Ventas y VentasLineas",
        "10. Pagos",
        "11. FulfillmentVentaLote",
        "12. MovimientosInventario (ajustes)"
    ]),
    ("CLAVES FORANEAS (FKs)", [
        "Las FKs referencian IDs de otras tablas.",
        "Validadas mediante listas (lst_*) definidas.",
        "Ejemplo: VentasLineas[venta_id] referencia Ventas[venta_id]",
        "Las listas se actualizan automáticamente cuando se agregan registros."
    ]),
    ("CATALOGOS", [
        "Listas de valores validos en hoja 'Catalogos'.",
        "Usados para campos tipo enum (estado, tipo, canal, etc).",
        "Validados mediante dropdown en las celdas.",
        "Catalogos: CAT_LoteEstado, CAT_MovTipo, CAT_PagoEstado, CAT_ClienteTipo, CAT_Canal, CAT_ItemTipo"
    ]),
    ("COLUMNAS CALCULADAS", [
        "Lotes[abv_estimado]: ABV basado en OG y FG medidos.",
        "Ventas[total_calc]: SUMIF de VentasLineas para la venta.",
        "Ventas[pagos_calc]: SUMIF de Pagos para la venta.",
        "Ventas[saldo_calc]: total_calc - pagos_calc."
    ])
]


def create_backup():
    """Create backup of existing file with timestamp"""
    if not OUTPUT_FILE.exists():
        print(f"No existing file to backup: {OUTPUT_FILE}")
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"Quelonio_Excel_MVP_Skeleton_step33_{timestamp}.xlsx"
    backup_path = BACKUP_DIR / backup_name

    shutil.copy2(OUTPUT_FILE, backup_path)
    print(f"Backup created: {backup_path}")
    return backup_path


def load_workbook_safe():
    """Load workbook with error handling"""
    if not OUTPUT_FILE.exists():
        raise FileNotFoundError(f"Excel file not found: {OUTPUT_FILE}")

    print(f"Loading workbook: {OUTPUT_FILE}")
    wb = load_workbook(OUTPUT_FILE, data_only=False)
    return wb


def create_home_sheet(wb):
    """Create 00_Home sheet as first sheet with navigation"""
    print("\nCreating 00_Home sheet...")

    sheet_name = "00_Home"

    # Remove if exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        print(f"  Removed existing {sheet_name} sheet")

    # Create as first sheet
    ws = wb.create_sheet(title=sheet_name, index=0)

    # Styles
    title_font = Font(size=20, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    link_font = Font(size=11, color="0563C1", underline="single")
    normal_font = Font(size=11)

    # Center alignment
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "QUELONIO - Sistema de Gestion Cervecera"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align

    # Section: Orden de Carga
    row = 3
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "ORDEN DE CARGA"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

    row += 1
    load_order = [
        "1. ItemsInventario (materias primas)",
        "2. Proveedores",
        "3. LotesInsumo (lotes de materias primas)",
        "4. Recetas y RecetaVersiones",
        "5. Clientes y Productos",
        "6. Lotes (produccion)",
        "7. Ventas, VentasLineas y Pagos",
        "8. FulfillmentVentaLote",
        "9. MovimientosInventario (ajustes)"
    ]

    for item in load_order:
        ws[f"A{row}"] = "•"
        ws[f"B{row}"] = item
        ws[f"B{row}"].font = normal_font
        ws[f"B{row}"].alignment = left_align
        row += 1

    row += 1

    # Section: Accesos
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = "ACCESOS RAPIDOS"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align

    row += 1
    for sheet_name in NAVIGATION_SHEETS:
        # Check if sheet exists
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet {sheet_name} not found, skipping link")
            continue

        # Create hyperlink
        link_formula = f'=HYPERLINK("#\'{sheet_name}\'!A1","{sheet_name}")'
        ws[f"A{row}"] = "→"
        ws[f"B{row}"].value = link_formula
        ws[f"B{row}"].font = link_font
        ws[f"B{row}"].alignment = left_align
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 50

    link_count = sum(1 for s in NAVIGATION_SHEETS if s in wb.sheetnames)
    print(f"  Created 00_Home as first sheet with {link_count} links")
    return ws


def create_ayuda_sheet(wb):
    """Create 98_Ayuda sheet with help content"""
    print("\nCreating 98_Ayuda sheet...")

    sheet_name = "98_Ayuda"

    # Remove if exists
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        print(f"  Removed existing {sheet_name} sheet")

    # Create sheet
    ws = wb.create_sheet(title=sheet_name)

    # Styles
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(size=14, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    normal_font = Font(size=11)

    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value = "AYUDA - Sistema de Gestion Cervecera"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center_align

    # Content sections
    row = 3
    for section_title, items in AYUDA_CONTENT:
        # Section header
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = section_title
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_align

        row += 1
        for item in items:
            ws[f"A{row}"] = ""
            ws[f"B{row}"] = item
            ws[f"B{row}"].font = normal_font
            ws[f"B{row}"].alignment = left_align
            row += 1

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 80

    print(f"  Created {sheet_name} with {len(AYUDA_CONTENT)} sections")
    return ws


def main():
    """Main function to build Excel MVP Step 33"""
    print("=" * 60)
    print("Quelonio Excel MVP Builder - Step 33")
    print("=" * 60)

    # Create backup
    backup_path = create_backup()

    # Load existing workbook
    try:
        wb = load_workbook_safe()
    except Exception as e:
        print(f"\nERROR: Failed to load workbook: {e}")
        return 1

    print(f"  Loaded {len(wb.sheetnames)} sheets")

    try:
        # Create Home sheet as first sheet
        create_home_sheet(wb)

        # Create Ayuda sheet
        create_ayuda_sheet(wb)

        # Save workbook
        wb.save(OUTPUT_FILE)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Excel MVP Step 33 completed:")
        print(f"  Path: {OUTPUT_FILE.absolute()}")
        print(f"  Sheets: {len(wb.sheetnames)}")
        print(f"  Backup: {backup_path}")
        print(f"{'=' * 60}\n")
        return 0

    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1
    finally:
        wb.close()


if __name__ == "__main__":
    exit(main())
