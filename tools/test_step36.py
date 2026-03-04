"""Tests for Excel MVP Step 36 - Dashboard with KPIs

Tests verify:
- 20_Dashboard sheet exists
- 5 KPIs exist with formulas
- stock_minimo column added if needed (table still valid)
- Compatibility with Step 35 (previous sheets/tables still exist)
"""

import subprocess
import sys
import os


def run_step35_tests():
    """Run Step 35 tests to verify backward compatibility."""
    print("\n" + "=" * 60)
    print("Running Step 35 compatibility tests...")
    print("=" * 60)
    
    step35_path = os.path.join("tools", "test_step35.py")
    if os.path.exists(step35_path):
        try:
            result = subprocess.run(
                [sys.executable, step35_path],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0:
                print("[OK] Step 35 tests PASSED")
                return True
            else:
                print("[FAIL] Step 35 tests FAILED")
                print(result.stdout)
                print(result.stderr)
                return False
        except subprocess.TimeoutExpired:
            print("[FAIL] Step 35 tests TIMED OUT")
            return False
        except Exception as e:
            print(f"[FAIL] Step 35 tests ERROR: {e}")
            return False
    else:
        print("[INFO] Step 35 test file not found (expected for new implementation)")
        return True


def test_dashboard_sheet_exists():
    """Test that 20_Dashboard sheet exists."""
    print("\n[Test] Verifying 20_Dashboard sheet exists...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheets = wb.sheetnames
        
        if "20_Dashboard" in sheets:
            print("[OK] 20_Dashboard sheet exists")
            wb.close()
            return True
        else:
            print(f"[FAIL] 20_Dashboard sheet not found. Available sheets: {sheets}")
            wb.close()
            return False
    except Exception as e:
        print(f"✗ Error checking dashboard sheet: {e}")
        return False


def test_kpis_exist():
    """Test that all 5 KPIs exist with non-empty formulas."""
    print("\n[Test] Verifying KPIs exist with formulas...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    expected_kpis = [
        "KPI_LotesEnCurso",
        "KPI_StockCriticoCount",
        "KPI_Ventas_30d",
        "KPI_SaldoPorCobrar",
        "KPI_MargenEstimado"
    ]
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["20_Dashboard"]
        
        found_kpis = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for kpi in expected_kpis:
                        if kpi in cell.value and kpi not in found_kpis:
                            found_kpis.append(kpi)
        
        # Check formulas for KPIs (based on actual KPI configuration in build script)
        kpi_formulas = {
            "KPI_LotesEnCurso": ws.cell(row=2, column=3).value,
            "KPI_StockCriticoCount": ws.cell(row=4, column=3).value,
            "KPI_Ventas_30d": ws.cell(row=6, column=3).value,
            "KPI_SaldoPorCobrar": ws.cell(row=8, column=3).value,
            "KPI_MargenEstimado": ws.cell(row=10, column=3).value,
        }
        
        all_passed = True
        for kpi, formula in kpi_formulas.items():
            if formula and isinstance(formula, str) and formula.startswith("="):
                print(f"  [OK] {kpi}: Formula present")
            else:
                print(f"  [FAIL] {kpi}: Formula missing or invalid: {formula}")
                all_passed = False
        
        # Check all KPI labels exist
        for kpi in expected_kpis:
            if kpi in found_kpis:
                print(f"  [OK] {kpi} label found")
            else:
                print(f"  [FAIL] {kpi} label not found")
                all_passed = False
        
        wb.close()
        return all_passed
    except Exception as e:
        print(f"[FAIL] Error checking KPIs: {e}")
        return False


def test_stock_minimo_column():
    """Test that stock_minimo column exists and table is valid."""
    print("\n[Test] Verifying stock_minimo column in 05_ItemsInventario...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["05_ItemsInventario"]
        
        # Check headers
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header).lower())
        
        if "stock_minimo" in headers:
            print("[OK] stock_minimo column exists")
            
            # Check column index
            col_idx = headers.index("stock_minimo") + 1
            print(f"  Column position: {col_idx}")
            
            # Verify it's not inflating table (should have data)
            has_data = False
            for row in range(2, min(10, ws.max_row + 1)):
                value = ws.cell(row=row, column=col_idx).value
                if value is not None:
                    has_data = True
                    break
            
            if has_data:
                print("  [OK] Column has data")
            else:
                print("  [WARN] Column has no data (acceptable if no items)")
            
            wb.close()
            return True
        else:
            print(f"[FAIL] stock_minimo column not found. Headers: {headers}")
            wb.close()
            return False
    except Exception as e:
        print(f"[FAIL] Error checking stock_minimo: {e}")
        return False


def test_previous_sheets_exist():
    """Test that previous Step 35 sheets still exist."""
    print("\n[Test] Verifying backward compatibility (Step 35 sheets)...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    expected_sheets = ["03_Lotes", "05_ItemsInventario", "11_Clientes", "12_Ventas"]
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheets = wb.sheetnames
        
        all_exist = True
        for sheet_name in expected_sheets:
            if sheet_name in sheets:
                print(f"  [OK] {sheet_name} exists")
            else:
                print(f"  [FAIL] {sheet_name} not found")
                all_exist = False
        
        wb.close()
        return all_exist
    except Exception as e:
        print(f"[FAIL] Error checking sheets: {e}")
        return False


def test_mini_panels():
    """Test that mini-panels exist on dashboard."""
    print("\n[Test] Verifying mini-panels on dashboard...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["20_Dashboard"]
        
        # Check for panel headers
        panel_headers = [
            "Lotes en Curso",
            "Stock Crítico",
            "Ventas y Cobranzas"
        ]
        
        all_found = True
        for header in panel_headers:
            found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and header in str(cell.value):
                        found = True
                        break
                if found:
                    break
            
            if found:
                print(f"  [OK] '{header}' panel found")
            else:
                print(f"  [FAIL] '{header}' panel not found")
                all_found = False
        
        wb.close()
        return all_found
    except Exception as e:
        print(f"[FAIL] Error checking mini-panels: {e}")
        return False


def run_all_tests():
    """Run all Step 36 tests."""
    print("=" * 60)
    print("Excel MVP Step 36 Tests - Dashboard with KPIs")
    print("=" * 60)
    
    # Test 1: Dashboard sheet exists
    test1 = test_dashboard_sheet_exists()
    
    # Test 2: Previous sheets exist (compatibility)
    test2 = test_previous_sheets_exist()
    
    # Test 3: KPIs exist with formulas
    test3 = test_kpis_exist()
    
    # Test 4: stock_minimo column exists
    test4 = test_stock_minimo_column()
    
    # Test 5: Mini-panels exist
    test5 = test_mini_panels()
    
    # Test 6: Run Step 35 tests (compatibility)
    test6 = run_step35_tests()
    
    # Summary
    print("\n" + "=" * 60)
    print("Test Summary")
    print("=" * 60)
    print(f"Dashboard sheet exists:         {'PASS' if test1 else 'FAIL'}")
    print(f"Previous sheets exist:           {'PASS' if test2 else 'FAIL'}")
    print(f"KPIs with formulas:             {'PASS' if test3 else 'FAIL'}")
    print(f"stock_minimo column:             {'PASS' if test4 else 'FAIL'}")
    print(f"Mini-panels exist:               {'PASS' if test5 else 'FAIL'}")
    print(f"Step 35 compatibility:           {'PASS' if test6 else 'FAIL'}")
    print("=" * 60)
    
    all_passed = all([test1, test2, test3, test4, test5, test6])
    
    if all_passed:
        print("[OK] ALL TESTS PASSED")
        return 0
    else:
        print("[FAIL] SOME TESTS FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(run_all_tests())
