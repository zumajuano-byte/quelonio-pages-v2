"""Export Excel tables to CSV pack with manifest

Exports table sheets (01_ to 15_) to CSV files in a timestamped folder,
generates pack_manifest.json with metadata and checksums.

Usage:
    python tools/export_excel_csv_pack.py [--delimiter comma|semicolon|tab] [--bom 0|1]
"""

import argparse
import csv
import hashlib
import json
import openpyxl
import os
from datetime import datetime
import sys
from datetime import datetime as dt


def get_table_sheets(wb):
    """Get list of table sheets (01_ to 15_)."""
    table_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith(('01_', '02_', '03_', '04_', '05_', '06_', '07_', '08_', '09_', '10_', '11_', '12_', '13_', '14_', '15_')):
            table_sheets.append(sheet_name)
    return sorted(table_sheets)


def get_delimiter(delimiter_str):
    """Convert delimiter string to character."""
    if delimiter_str == 'comma':
        return ','
    elif delimiter_str == 'semicolon':
        return ';'
    elif delimiter_str == 'tab':
        return '\t'
    else:
        raise ValueError(f"Invalid delimiter: {delimiter_str}")


def export_sheet_to_csv(ws, csv_path, delimiter, bom):
    """Export worksheet to CSV with specified delimiter and BOM."""
    # Get all data, format dates
    data = []
    for row in ws.iter_rows(values_only=True):
        formatted_row = []
        for cell in row:
            if isinstance(cell, datetime):
                formatted_row.append(cell.strftime('%Y-%m-%d'))
            else:
                formatted_row.append(cell)
        data.append(formatted_row)

    # Write CSV
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        if bom:
            f.write('\ufeff')  # UTF-8 BOM
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerows(data)

    return len(data)


def calculate_sha256(file_path):
    """Calculate SHA256 of file bytes."""
    hash_sha256 = hashlib.sha256()
    with open(file_path, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_sha256.update(chunk)
    return hash_sha256.hexdigest()


def create_manifest(export_folder, workbook_path, delimiter, bom, files_info):
    """Create pack_manifest.json."""
    manifest = {
        "schema_version": "mvp-v1",
        "created_at": datetime.now().isoformat(),
        "delimiter": delimiter,
        "bom": bom,
        "encoding": "utf-8",
        "workbook_source": workbook_path,
        "files": files_info,
        "total_files": len(files_info),
        "notes": "Quelonio Excel MVP CSV Pack"
    }

    manifest_path = os.path.join(export_folder, 'pack_manifest.json')
    with open(manifest_path, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    return manifest_path


def export_csv_pack(workbook_path, delimiter_str='semicolon', bom=0):
    """Main export function."""
    print("Exporting Excel CSV Pack...")
    print(f"Workbook: {workbook_path}")

    # Validate workbook exists
    if not os.path.exists(workbook_path):
        print(f"[ERROR] Workbook not found: {workbook_path}")
        return None

    # Load workbook
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as e:
        print(f"[ERROR] Failed to load workbook: {e}")
        return None

    # Get table sheets
    table_sheets = get_table_sheets(wb)
    if not table_sheets:
        print("[WARNING] No table sheets found (01_ to 15_)")
        wb.close()
        return None

    print(f"Found {len(table_sheets)} table sheets: {', '.join(table_sheets)}")

    # Create export folder
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_folder = os.path.join("data", "excel", "exports", f"csv_pack_{timestamp}")
    os.makedirs(export_folder, exist_ok=True)
    print(f"Export folder: {export_folder}")

    # Export settings
    delimiter = get_delimiter(delimiter_str)
    bom = bool(bom)
    print(f"Delimiter: {delimiter_str} ({repr(delimiter)})")
    print(f"BOM: {bom}")

    # Export each sheet to CSV
    files_info = []
    for sheet_name in table_sheets:
        ws = wb[sheet_name]
        csv_filename = f"{sheet_name}.csv"
        csv_path = os.path.join(export_folder, csv_filename)

        # Export to CSV
        rows = export_sheet_to_csv(ws, csv_path, delimiter, bom)
        print(f"  Exported {sheet_name} -> {csv_filename} ({rows} rows)")

        # Calculate checksum
        sha256 = calculate_sha256(csv_path)

        files_info.append({
            "name": csv_filename,
            "rows": rows,
            "sha256": sha256
        })

    wb.close()

    # Create manifest
    manifest_path = create_manifest(export_folder, workbook_path, delimiter_str, bom, files_info)
    print(f"Manifest: {manifest_path}")

    print(f"[OK] Export complete: {len(table_sheets)} CSV files + manifest")
    return export_folder


def main():
    parser = argparse.ArgumentParser(description="Export Excel tables to CSV pack with manifest")
    parser.add_argument('--delimiter', choices=['comma', 'semicolon', 'tab'], default='semicolon',
                       help='CSV delimiter (default: semicolon)')
    parser.add_argument('--bom', type=int, choices=[0, 1], default=0,
                       help='Include UTF-8 BOM (default: 0)')

    args = parser.parse_args()

    workbook_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")

    export_folder = export_csv_pack(workbook_path, args.delimiter, args.bom)
    if export_folder:
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()