#!/usr/bin/env python
"""
Test script for Excel MVP Step 32
Verifica: Catalogos, Defined Names, Data Validations, Calculated Columns
"""

import os
from pathlib import Path
from openpyxl import load_workbook

# Configuration
OUTPUT_FILE = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")

# Expected catalog lists
EXPECTED_CATALOGS = [
    "CAT_LoteEstado",
    "CAT_MovTipo",
    "CAT_PagoEstado",
    "CAT_ClienteTipo",
    "CAT_Canal",
    "CAT_ItemTipo"
]

# Expected catalog defined names
EXPECTED_CATALOG_DEFINED_NAMES = [
    "lst_lote_estado",
    "lst_mov_tipo",
    "lst_pago_estado",
    "lst_cliente_tipo",
    "lst_canal",
    "lst_item_tipo"
]

# Expected ID defined names (FK targets)
EXPECTED_ID_DEFINED_NAMES = [
    "lst_receta_id",
    "lst_recipe_version_id",
    "lst_batch_id",
    "lst_item_id",
    "lst_lote_insumo_id",
    "lst_venta_id",
    "lst_linea_id",
    "lst_producto_id"
]

# Expected data validation columns (must have validation)
EXPECTED_VALIDATIONS = [
    ("03_Lotes", "estado"),
    ("03_Lotes", "recipe_version_id"),
    ("06_MovimientosInventario", "tipo_movimiento"),
    ("13_VentasLineas", "venta_id"),
    ("13_VentasLineas", "producto_id")
]

# Expected calculated columns
EXPECTED_CALCULATED_COLUMNS = {
    "03_Lotes": ["abv_estimado"],
    "12_Ventas": ["total_calc", "pagos_calc", "saldo_calc"]
}


def test_file_exists():
    """Test 1: Verify the Excel file exists"""
    print("\nTest 1: File Existence")
    print("-" * 50)

    if not OUTPUT_FILE.exists():
        print(f"[X] FAIL: File does not exist: {OUTPUT_FILE}")
        return False

    print(f"[OK] PASS: File exists")
    return True


def test_catalogos_sheet_exists():
    """Test 2: Verify Catalogos sheet exists"""
    print("\nTest 2: Catalogos Sheet Existence")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        if "Catalogos" not in wb.sheetnames:
            print(f"[X] FAIL: Catalogos sheet not found")
            print(f"   Available sheets: {wb.sheetnames}")
            wb.close()
            return False

        ws = wb["Catalogos"]

        # Check that catalogs are present
        catalog_names = []
        for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
            if row[0]:
                catalog_names.append(row[0])

        missing_catalogs = []
        for catalog in EXPECTED_CATALOGS:
            if catalog not in catalog_names:
                missing_catalogs.append(catalog)

        if missing_catalogs:
            print(f"[X] FAIL: Missing catalogs in Catalogos sheet")
            print(f"   Missing: {missing_catalogs}")
            print(f"   Found: {catalog_names}")
            wb.close()
            return False

        print(f"[OK] PASS: Catalogos sheet exists with all catalogs")
        print(f"   Catalogs: {', '.join(EXPECTED_CATALOGS)}")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking Catalogos sheet: {e}")
        return False


def test_catalog_defined_names():
    """Test 3: Verify catalog defined names exist"""
    print("\nTest 3: Catalog Defined Names")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_names = list(wb.defined_names.keys())

        missing_names = []
        for expected_name in EXPECTED_CATALOG_DEFINED_NAMES:
            if expected_name not in actual_names:
                missing_names.append(expected_name)

        if missing_names:
            print(f"[X] FAIL: Missing catalog defined names")
            print(f"   Missing: {missing_names}")
            print(f"   Expected: {EXPECTED_CATALOG_DEFINED_NAMES}")
            wb.close()
            return False

        print(f"[OK] PASS: All catalog defined names present ({len(EXPECTED_CATALOG_DEFINED_NAMES)})")
        print(f"   Names: {', '.join(EXPECTED_CATALOG_DEFINED_NAMES)}")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking catalog defined names: {e}")
        return False


def test_id_defined_names():
    """Test 4: Verify ID defined names (FK targets) exist"""
    print("\nTest 4: ID Defined Names (FK Targets)")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        actual_names = list(wb.defined_names.keys())

        missing_names = []
        for expected_name in EXPECTED_ID_DEFINED_NAMES:
            if expected_name not in actual_names:
                missing_names.append(expected_name)

        if missing_names:
            print(f"[X] FAIL: Missing ID defined names")
            print(f"   Missing: {missing_names}")
            print(f"   Expected: {EXPECTED_ID_DEFINED_NAMES}")
            wb.close()
            return False

        print(f"[OK] PASS: All ID defined names present ({len(EXPECTED_ID_DEFINED_NAMES)})")
        print(f"   Names: {', '.join(EXPECTED_ID_DEFINED_NAMES)}")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking ID defined names: {e}")
        return False


def test_data_validations_exist():
    """Test 5: Verify data validations exist in expected columns"""
    print("\nTest 5: Data Validations in Expected Columns")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        all_pass = True

        for sheet_name, column_name in EXPECTED_VALIDATIONS:
            if sheet_name not in wb.sheetnames:
                print(f"[X] FAIL: Sheet '{sheet_name}' not found")
                all_pass = False
                continue

            ws = wb[sheet_name]

            # Find column index
            col_idx = None
            for idx, cell in enumerate(ws[1], start=0):
                if cell.value == column_name:
                    col_idx = idx
                    break

            if col_idx is None:
                print(f"[X] FAIL: Column '{column_name}' not found in '{sheet_name}'")
                all_pass = False
                continue

            # Check if column has data validation
            col_letter = chr(ord('A') + col_idx)
            has_validation = False

            # Access dataValidation list from DataValidationList object
            if hasattr(ws.data_validations, 'dataValidation'):
                dv_list = ws.data_validations.dataValidation
            else:
                # Fallback: iterate directly over data_validations
                dv_list = ws.data_validations

            for dv in dv_list:
                # Check sqref attribute (contains MultiCellRange)
                if hasattr(dv, 'sqref') and dv.sqref:
                    sqref_str = str(dv.sqref)
                    if sqref_str.startswith(col_letter):
                        has_validation = True
                        break

            if not has_validation:
                print(f"[X] FAIL: No data validation in {sheet_name}[{column_name}]")
                all_pass = False
            else:
                print(f"[OK] PASS: Data validation found in {sheet_name}[{column_name}]")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking data validations: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_calculated_columns_formulas():
    """Test 6: Verify calculated columns exist with correct formulas"""
    print("\nTest 6: Calculated Columns and Formulas")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        all_pass = True

        for sheet_name, expected_columns in EXPECTED_CALCULATED_COLUMNS.items():
            if sheet_name not in wb.sheetnames:
                print(f"[X] FAIL: Sheet '{sheet_name}' not found")
                all_pass = False
                continue

            ws = wb[sheet_name]

            for column_name in expected_columns:
                # Find column index
                col_idx = None
                for idx, cell in enumerate(ws[1], start=0):
                    if cell.value == column_name:
                        col_idx = idx
                        break

                if col_idx is None:
                    print(f"[X] FAIL: Column '{column_name}' not found in '{sheet_name}'")
                    all_pass = False
                    continue

                # Get formula from row 2
                cell = ws.cell(row=2, column=col_idx + 1)
                formula = cell.value

                if not formula or not isinstance(formula, str):
                    print(f"[X] FAIL: No formula in {sheet_name}[{column_name}]")
                    all_pass = False
                    continue

                # Check specific formula content
                if sheet_name == "03_Lotes" and column_name == "abv_estimado":
                    if "131.25" not in formula:
                        print(f"[X] FAIL: Formula in {sheet_name}[{column_name}] does not contain '131.25'")
                        print(f"   Formula: {formula}")
                        all_pass = False
                    else:
                        print(f"[OK] PASS: Formula in {sheet_name}[{column_name}] contains '131.25'")

                elif sheet_name == "12_Ventas" and column_name in ["total_calc", "pagos_calc"]:
                    if "SUMIF" not in formula:
                        print(f"[X] FAIL: Formula in {sheet_name}[{column_name}] does not contain 'SUMIF'")
                        print(f"   Formula: {formula}")
                        all_pass = False
                    else:
                        print(f"[OK] PASS: Formula in {sheet_name}[{column_name}] contains 'SUMIF'")

                elif sheet_name == "12_Ventas" and column_name == "saldo_calc":
                    if "total_calc" not in formula or "pagos_calc" not in formula:
                        print(f"[X] FAIL: Formula in {sheet_name}[{column_name}] does not reference calculated columns")
                        print(f"   Formula: {formula}")
                        all_pass = False
                    else:
                        print(f"[OK] PASS: Formula in {sheet_name}[{column_name}] references calculated columns")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking calculated columns: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_step31_not_broken():
    """Test 7: Verify Step 31 tests still pass"""
    print("\nTest 7: Step 31 Compatibility")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        # Check that all Step 31 sheets still exist (plus Catalogos)
        expected_sheets_step31 = [
            "01_Recetas",
            "02_RecetaVersiones",
            "03_Lotes",
            "04_LoteMediciones",
            "05_ItemsInventario",
            "06_MovimientosInventario",
            "07_Proveedores",
            "08_LotesInsumo",
            "09_ConsumosLote",
            "10_Productos",
            "11_Clientes",
            "12_Ventas",
            "13_VentasLineas",
            "14_Pagos",
            "15_FulfillmentVentaLote",
            "99_Listas",
            "Catalogos"
        ]

        actual_sheets = wb.sheetnames

        missing_sheets = []
        for sheet in expected_sheets_step31:
            if sheet not in actual_sheets:
                missing_sheets.append(sheet)

        if missing_sheets:
            print(f"[X] FAIL: Missing expected sheets (Step 31 compatibility)")
            print(f"   Missing: {missing_sheets}")
            wb.close()
            return False

        # Check Step 31 defined names still exist
        step31_defined_names = [
            "LIST_RECETAS_ESTADO",
            "LIST_LOTES_ESTADO",
            "LIST_ITEMS_TIPO",
            "LIST_ITEMS_UNIDAD_MEDIDA",
            "LIST_ITEMS_ESTADO",
            "LIST_MOVIMIENTOS_TIPO",
            "LIST_CLIENTES_TIPO",
            "LIST_CLIENTES_CANAL",
            "LIST_VENTAS_ESTADO_PAGO",
            "LIST_PAGOS_METODO",
            "LIST_PRODUCTOS_ESTADO"
        ]

        actual_names = list(wb.defined_names.keys())

        missing_names = []
        for name in step31_defined_names:
            if name not in actual_names:
                missing_names.append(name)

        if missing_names:
            print(f"[X] FAIL: Missing Step 31 defined names")
            print(f"   Missing: {missing_names}")
            wb.close()
            return False

        print(f"[OK] PASS: Step 31 sheets and defined names preserved")
        print(f"   Sheets: {len(actual_sheets)}")
        print(f"   Step 31 names: {len(step31_defined_names)}")
        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking Step 31 compatibility: {e}")
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel MVP Tests - Step 32")
    print("=" * 60)

    # Run all tests
    results = []
    results.append(("File Exists", test_file_exists()))
    results.append(("Catalogos Sheet", test_catalogos_sheet_exists()))
    results.append(("Catalog Defined Names", test_catalog_defined_names()))
    results.append(("ID Defined Names", test_id_defined_names()))
    results.append(("Data Validations", test_data_validations_exist()))
    results.append(("Calculated Columns", test_calculated_columns_formulas()))
    results.append(("Step 31 Compatible", test_step31_not_broken()))

    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] PASS" if result else "[X] FAIL"
        print(f"{status}: {test_name}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n*** All tests passed! ***")
        return 0
    else:
        print(f"\n*** {total - passed} test(s) failed ***")
        return 1


if __name__ == "__main__":
    exit(main())
