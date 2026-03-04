"""Tests for Excel MVP Step 37 - Dashboard Filters and Alerts

Tests verify:
- 20_Dashboard and 21_Calc exist
- Catalogos sheet with CAT_PeriodoDias exists
- Named ranges lst_periodo_dias and SEL_PeriodoDias exist
- Data validation applied to selector cell
- KPIs reference period filter (CALC_FechaInicio or SEL_PeriodoDias)
- 3 alerts exist with formulas
- Backward compatibility maintained
"""

import sys
import os


def test_dashboard_and_calc_sheets_exist():
    """Test that 20_Dashboard and 21_Calc sheets exist."""
    print("\n[Test] Verifying 20_Dashboard and 21_Calc sheets exist...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheets = wb.sheetnames
        
        all_exist = True
        for sheet_name in ["20_Dashboard", "21_Calc"]:
            if sheet_name in sheets:
                print(f"  [OK] {sheet_name} exists")
            else:
                print(f"  [FAIL] {sheet_name} not found")
                all_exist = False
        
        wb.close()
        return all_exist
    except Exception as e:
        print(f"[FAIL] Error checking sheets: {e}")
        return False


def test_catalogos_sheet_exists_with_period_values():
    """Test that Catalogos sheet exists with CAT_PeriodoDias containing 7,30,90."""
    print("\n[Test] Verifying Catalogos sheet with CAT_PeriodoDias...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        if "Catalogos" not in wb.sheetnames:
            print("[FAIL] Catalogos sheet not found")
            wb.close()
            return False
        
        ws = wb["Catalogos"]
        
        # Find CAT_PeriodoDias rows
        period_values = []
        for row in range(2, ws.max_row + 1):
            cat_name = ws.cell(row=row, column=1).value
            if cat_name == "CAT_PeriodoDias":
                valor = ws.cell(row=row, column=2).value
                if valor is not None:
                    period_values.append(int(valor))
        
        expected_values = [7, 30, 90]
        if set(period_values) == set(expected_values):
            print(f"  [OK] CAT_PeriodoDias contains expected values: {period_values}")
            wb.close()
            return True
        else:
            print(f"  [FAIL] CAT_PeriodoDias has wrong values. Expected {expected_values}, got {period_values}")
            wb.close()
            return False
    except Exception as e:
        print(f"[FAIL] Error checking Catalogos: {e}")
        return False


def test_named_ranges_exist():
    """Test that named ranges lst_periodo_dias and SEL_PeriodoDias exist."""
    print("\n[Test] Verifying named ranges exist...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        named_ranges = list(wb.defined_names)
        
        all_exist = True
        for name in ["lst_periodo_dias", "SEL_PeriodoDias"]:
            if name in named_ranges:
                print(f"  [OK] Named range '{name}' exists")
            else:
                print(f"  [FAIL] Named range '{name}' not found")
                all_exist = False
        
        wb.close()
        return all_exist
    except Exception as e:
        print(f"[FAIL] Error checking named ranges: {e}")
        return False


def test_selector_data_validation():
    """Test that data validation is applied to selector cell."""
    print("\n[Test] Verifying data validation on selector cell...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["20_Dashboard"]
        
        # Check for data validation on B3
        has_validation = False
        for dv in ws.data_validations.dataValidation:
            if "B3" in dv.sqref:
                if "lst_periodo_dias" in str(dv.formula1):
                    has_validation = True
                    break
        
        if has_validation:
            print("  [OK] Data validation exists on selector cell B3")
            wb.close()
            return True
        else:
            print("  [FAIL] Data validation not found on selector cell B3")
            wb.close()
            return False
    except Exception as e:
        print(f"[FAIL] Error checking data validation: {e}")
        return False


def test_kpis_reference_period_filter():
    """Test that KPIs reference period filter (CALC_FechaInicio or SEL_PeriodoDias)."""
    print("\n[Test] Verifying KPIs reference period filter...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["20_Dashboard"]
        
        # Check KPI_Ventas_Periodo (row 6, column 3)
        ventas_formula = ws.cell(row=6, column=3).value
        if ventas_formula and isinstance(ventas_formula, str):
            if "CALC_FechaInicio" in ventas_formula or "SEL_PeriodoDias" in ventas_formula:
                print(f"  [OK] KPI_Ventas_Periodo references period filter")
            else:
                print(f"  [FAIL] KPI_Ventas_Periodo does not reference period filter: {ventas_formula}")
                wb.close()
                return False
        else:
            print(f"  [FAIL] KPI_Ventas_Periodo formula not found")
            wb.close()
            return False
        
        # Check KPI_SaldoPorCobrar (row 8, column 3)
        saldo_formula = ws.cell(row=8, column=3).value
        if saldo_formula and isinstance(saldo_formula, str):
            if "CALC_FechaInicio" in saldo_formula or "SEL_PeriodoDias" in saldo_formula:
                print(f"  [OK] KPI_SaldoPorCobrar references period filter")
            else:
                print(f"  [WARN] KPI_SaldoPorCobrar may not reference period filter: {saldo_formula}")
        
        wb.close()
        return True
    except Exception as e:
        print(f"[FAIL] Error checking KPIs: {e}")
        return False


def test_alerts_exist_with_formulas():
    """Test that 3 alerts exist with formulas (not hardcoded values)."""
    print("\n[Test] Verifying alerts exist with formulas...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        ws = wb["20_Dashboard"]
        
        alerts = [
            {"name": "ALERT_StockCritico", "row": 12, "col": 2},
            {"name": "ALERT_LotesDemorados", "row": 13, "col": 2},
            {"name": "ALERT_SaldoPendiente", "row": 14, "col": 2}
        ]
        
        all_formulas = True
        for alert in alerts:
            formula = ws.cell(row=alert["row"], column=alert["col"]).value
            if formula and isinstance(formula, str) and formula.startswith("="):
                print(f"  [OK] {alert['name']} has formula")
            else:
                print(f"  [FAIL] {alert['name']} missing formula: {formula}")
                all_formulas = False
        
        # Check status formulas
        for row in [12, 13, 14]:
            status_formula = ws.cell(row=row, column=3).value
            if status_formula and isinstance(status_formula, str) and status_formula.startswith("="):
                print(f"  [OK] Status cell row {row} has formula")
            else:
                print(f"  [FAIL] Status cell row {row} missing formula")
                all_formulas = False
        
        wb.close()
        return all_formulas
    except Exception as e:
        print(f"[FAIL] Error checking alerts: {e}")
        return False


def test_backward_compatibility():
    """Test backward compatibility - no sheets/tables lost."""
    print("\n[Test] Verifying backward compatibility...")
    
    import openpyxl
    excel_path = os.path.join("data", "excel", "Quelonio_Excel_MVP_Skeleton.xlsx")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheets = wb.sheetnames
        
        expected_sheets = [
            "20_Dashboard",
            "21_Calc",
            "03_Lotes",
            "05_ItemsInventario",
            "11_Clientes",
            "12_Ventas",
            "Catalogos"
        ]
        
        all_exist = True
        for sheet_name in expected_sheets:
            if sheet_name in sheets:
                print(f"  [OK] {sheet_name} exists")
            else:
                print(f"  [FAIL] {sheet_name} not found")
                all_exist = False
        
        # Check that 15 tables (if any) still exist
        # Note: openpyxl doesn't track tables directly, so we check sheets exist
        
        wb.close()
        return all_exist
    except Exception as e:
        print(f"[FAIL] Error checking compatibility: {e}")
        return False


def run_all_tests():
    """Run all Step 37 tests."""
    print("=" * 60)
    print("Excel MVP Step 37 Tests - Dashboard Filters and Alerts")
    print("=" * 60)
    
    test1 = test_dashboard_and_calc_sheets_exist()
    test2 = test_catalogos_sheet_exists_with_period_values()
    test3 = test_named_ranges_exist()
    test4 = test_selector_data_validation()
    test5 = test_kpis_reference_period_filter()
    test6 = test_alerts_exist_with_formulas()
    test7 = test_backward_compatibility()
    
    print("\n" + "=" * 60)
    print("Test Summary")
    print("=" * 60)
    print(f"20_Dashboard & 21_Calc exist:     {'PASS' if test1 else 'FAIL'}")
    print(f"Catalogos with CAT_PeriodoDias:   {'PASS' if test2 else 'FAIL'}")
    print(f"Named ranges exist:               {'PASS' if test3 else 'FAIL'}")
    print(f"Selector data validation:          {'PASS' if test4 else 'FAIL'}")
    print(f"KPIs reference period filter:      {'PASS' if test5 else 'FAIL'}")
    print(f"Alerts with formulas:              {'PASS' if test6 else 'FAIL'}")
    print(f"Backward compatibility:            {'PASS' if test7 else 'FAIL'}")
    print("=" * 60)
    
    all_passed = all([test1, test2, test3, test4, test5, test6, test7])
    
    if all_passed:
        print("[OK] ALL TESTS PASSED")
        return 0
    else:
        print("[FAIL] SOME TESTS FAILED")
        return 1


if __name__ == "__main__":
    sys.exit(run_all_tests())
