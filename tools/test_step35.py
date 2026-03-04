#!/usr/bin/env python
"""
Test script for Excel MVP Step 35
Verifica: Conteos reales, rangos de tablas, export CSV limpio
"""

from pathlib import Path
from openpyxl import load_workbook
import re

# Configuration
OUTPUT_FILE = Path("data/excel/Quelonio_Excel_MVP_Skeleton.xlsx")
EXPORT_DIR = Path("data/excel/exports")

# Expected real row counts (based on seed data)
EXPECTED_ROW_COUNTS = {
    "01_Recetas": 2,
    "02_RecetaVersiones": 3,
    "03_Lotes": 2,  # KEY: Should be 2, not 5001
    "04_LoteMediciones": 0,
    "05_ItemsInventario": 10,
    "06_MovimientosInventario": 0,
    "07_Proveedores": 2,
    "08_LotesInsumo": 5,
    "09_ConsumosLote": 10,
    "10_Productos": 2,
    "11_Clientes": 3,
    "12_Ventas": 3,  # KEY: Should be 3, not 5002
    "13_VentasLineas": 5,
    "14_Pagos": 3,
    "15_FulfillmentVentaLote": 5
}

# ID column for each sheet
SHEET_ID_COLUMNS = {
    "01_Recetas": "receta_id",
    "02_RecetaVersiones": "recipe_version_id",
    "03_Lotes": "batch_id",
    "04_LoteMediciones": "medicion_id",
    "05_ItemsInventario": "item_id",
    "06_MovimientosInventario": "movimiento_id",
    "07_Proveedores": "proveedor_id",
    "08_LotesInsumo": "lote_insumo_id",
    "09_ConsumosLote": "consumo_id",
    "10_Productos": "producto_id",
    "11_Clientes": "cliente_id",
    "12_Ventas": "venta_id",
    "13_VentasLineas": "linea_id",
    "14_Pagos": "pago_id",
    "15_FulfillmentVentaLote": "fulfillment_id"
}


def test_file_exists():
    """Test 1: Verify Excel file exists"""
    print("\nTest 1: File Existence")
    print("-" * 50)

    if not OUTPUT_FILE.exists():
        print(f"[X] FAIL: File does not exist: {OUTPUT_FILE}")
        return False

    print(f"[OK] PASS: File exists")
    return True


def test_real_row_counts():
    """Test 2: Verify real row counts based on IDs (wherever they are)"""
    print("\nTest 2: Real Row Counts (Key Fix)")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=True)
        all_pass = True

        # Critical sheets that were inflated
        critical_sheets = ["03_Lotes", "12_Ventas"]

        for sheet_name, expected_count in EXPECTED_ROW_COUNTS.items():
            if sheet_name not in wb.sheetnames:
                print(f"[X] FAIL: Sheet '{sheet_name}' not found")
                all_pass = False
                continue

            ws = wb[sheet_name]

            # Count actual rows with non-empty IDs (wherever they are)
            id_column = SHEET_ID_COLUMNS.get(sheet_name)

            if not id_column:
                print(f"  WARNING: No ID column for {sheet_name}")
                continue

            # Find ID column
            header_map = {}
            for col_idx, cell in enumerate(ws[1], start=0):
                if cell.value:
                    header_map[cell.value] = col_idx

            if id_column not in header_map:
                print(f"  WARNING: ID column '{id_column}' not found in {sheet_name}")
                continue

            col_idx = header_map[id_column]

            # Count rows with non-empty IDs
            real_count = 0
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx + 1).value
                if cell_value is not None and cell_value != "":
                    real_count += 1

            if real_count != expected_count:
                if sheet_name in critical_sheets:
                    print(f"[X] FAIL: {sheet_name} has {real_count} rows with IDs, expected {expected_count}")
                    all_pass = False
                else:
                    print(f"[X] FAIL: {sheet_name} has {real_count} rows with IDs, expected {expected_count}")
                    all_pass = False
            else:
                status = "KEY FIX" if sheet_name in critical_sheets else "OK"
                print(f"[{status}] PASS: {sheet_name} has {real_count} rows with IDs (expected {expected_count})")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking row counts: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_table_references_not_inflated():
    """Test 3: Verify table references are NOT inflated - strict check"""
    print("\nTest 3: Table References Not Inflated (STRICT)")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)
        all_pass = True

        for sheet_name, expected_count in EXPECTED_ROW_COUNTS.items():
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]

            # Check if table exists
            if not hasattr(ws, 'tables') or len(ws.tables) == 0:
                print(f"  INFO: {sheet_name} - No table found")
                continue

            table_name = list(ws.tables.keys())[0]
            table = ws.tables[table_name]

            # Parse table reference
            table_ref = table.ref

            # Extract end row from table ref
            match = re.search(r':([A-Z]+)(\d+)$', table_ref)
            if match:
                end_row = int(match.group(2))

                # Expected end row is expected_count + 1 (for header)
                expected_end_row = expected_count + 1

                # STRICT CHECK: Table ref should be at most 5 rows above expected
                # This means data is at correct position, not at row 5000+
                max_allowed_end_row = expected_end_row + 5

                if end_row > max_allowed_end_row:
                    print(f"[X] FAIL: {sheet_name} table ref ends at row {end_row}, expected ~{expected_end_row}")
                    print(f"  Table ref: {table_ref}")
                    print(f"  Max allowed: {max_allowed_end_row} (expected + 5)")
                    all_pass = False
                elif end_row < expected_end_row:
                    print(f"[X] FAIL: {sheet_name} table ref ends at row {end_row}, expected {expected_end_row}")
                    print(f"  Table ref: {table_ref}")
                    print(f"  Table is too small!")
                    all_pass = False
                else:
                    print(f"[OK] PASS: {sheet_name} table ref: {table_ref} (reasonable)")
            else:
                print(f"  WARNING: Could not parse table ref for {sheet_name}: {table_ref}")

        wb.close()
        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking table references: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_csv_export_clean():
    """Test 4: Verify CSV exports have reasonable line counts (only real data)"""
    print("\nTest 4: CSV Export Clean (No Inflated Data)")
    print("-" * 50)

    try:
        # Find latest export folder
        if not EXPORT_DIR.exists():
            print(f"[X] FAIL: Export directory not found: {EXPORT_DIR}")
            return False

        export_folders = sorted([d for d in EXPORT_DIR.iterdir() if d.is_dir() and d.name.startswith("STEP34_pack_") or d.name.startswith("STEP35_pack_")])

        if not export_folders:
            print(f"[X] FAIL: No export folders found in {EXPORT_DIR}")
            return False

        latest_folder = export_folders[-1]

        all_pass = True

        # Check critical files - count non-empty data rows
        critical_files = {
            "03_Lotes.csv": 2,  # 1 header + 2 data rows
            "12_Ventas.csv": 3   # 1 header + 3 data rows
        }

        for csv_name, expected_lines in critical_files.items():
            csv_path = latest_folder / csv_name

            if not csv_path.exists():
                print(f"[X] FAIL: CSV file not found: {csv_name}")
                all_pass = False
                continue

            # Count non-empty data lines (excluding header)
            import csv
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=';')
                rows = list(reader)

            if not rows:
                print(f"[X] FAIL: {csv_name} is empty")
                all_pass = False
                continue

            data_rows = [row for row in rows[1:] if any(cell.strip() for cell in row)]  # Skip header, exclude empty rows
            actual_data_count = len(data_rows)

            if actual_data_count != expected_lines:
                print(f"[X] FAIL: {csv_name} has {actual_data_count} non-empty data rows, expected {expected_lines}")
                all_pass = False
            else:
                total_lines = len(rows)  # header + data
                print(f"[OK] PASS: {csv_name} has {actual_data_count} data rows (total {total_lines} lines)")

        return all_pass

    except Exception as e:
        print(f"[X] FAIL: Error checking CSV exports: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_steps_31_34_compatibility():
    """Test 5: Verify Steps 31-34 compatibility"""
    print("\nTest 5: Steps 31-34 Compatibility")
    print("-" * 50)

    try:
        wb = load_workbook(OUTPUT_FILE, data_only=False)

        # Check for Step 31 sheets
        step31_sheets = [
            "01_Recetas",
            "02_RecetaVersiones",
            "03_Lotes",
            "04_LoteMediciones",
            "05_ItemsInventario",
            "06_MovimientosInventario",
            "07_Proveedores",
            "08_LotesInsumo",
            "09_ConsumosLote",
            "10_Productos",
            "11_Clientes",
            "12_Ventas",
            "13_VentasLineas",
            "14_Pagos",
            "15_FulfillmentVentaLote",
            "99_Listas"
        ]

        missing_sheets = []
        for sheet in step31_sheets:
            if sheet not in wb.sheetnames:
                missing_sheets.append(sheet)

        if missing_sheets:
            print(f"[X] FAIL: Missing Step 31 sheets")
            wb.close()
            return False

        # Check for Step 32 catalog sheet
        if "Catalogos" not in wb.sheetnames:
            print(f"[X] FAIL: Catalogos sheet not found (Step 32)")
            wb.close()
            return False

        # Check for Step 33 sheets
        step33_sheets = ["00_Home", "98_Ayuda"]
        for sheet in step33_sheets:
            if sheet not in wb.sheetnames:
                print(f"[X] FAIL: Step 33 sheet missing: {sheet}")
                wb.close()
                return False

        # Check that seed data still exists (Step 34)
        ws_recetas = wb["01_Recetas"]
        if ws_recetas.max_row < 2:
            print(f"[X] FAIL: Seed data from Step 34 missing (01_Recetas)")
            wb.close()
            return False

        print(f"[OK] PASS: Steps 31-34 compatible")
        print(f"   Step 31 sheets: {len(step31_sheets)}")
        print(f"   Step 32 Catalogos: OK")
        print(f"   Step 33 sheets: {len(step33_sheets)}")
        print(f"   Step 34 seed data: OK")

        wb.close()
        return True

    except Exception as e:
        print(f"[X] FAIL: Error checking compatibility: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Run all tests"""
    print("=" * 60)
    print("Excel MVP Tests - Step 35 (Hygiene)")
    print("=" * 60)

    # Run all tests
    results = []
    results.append(("File Exists", test_file_exists()))
    results.append(("Real Row Counts", test_real_row_counts()))
    results.append(("Table References", test_table_references_not_inflated()))
    results.append(("CSV Export Clean", test_csv_export_clean()))
    results.append(("Steps 31-34 Compatibility", test_steps_31_34_compatibility()))

    # Print summary
    print("\n" + "=" * 60)
    print("TEST SUMMARY")
    print("=" * 60)

    passed = sum(1 for _, result in results if result)
    total = len(results)

    for test_name, result in results:
        status = "[OK] PASS" if result else "[X] FAIL"
        print(f"{status}: {test_name}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n*** All tests passed! ***")
        return 0
    else:
        print(f"\n*** {total - passed} test(s) failed ***")
        return 1


if __name__ == "__main__":
    exit(main())
